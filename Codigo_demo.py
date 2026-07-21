#!/usr/bin/env python3
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import time
import collections
import sqlite3
import os
import json
import random
import math
import subprocess
from datetime import datetime

# --- CONFIGURACIÓN DE GRÁFICA ---
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

# --- IMPORTACIONES PARA LOGOS (PIL) ---
try:
    from PIL import Image, ImageTk
    PIL_DISPONIBLE = True
    try:
        RESAMPLE_METHOD = Image.LANCZOS
    except AttributeError:
        RESAMPLE_METHOD = Image.ANTIALIAS
except ImportError:
    PIL_DISPONIBLE = False

# --- RUTAS DE ARCHIVOS ---
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ASSETS_DIR = os.path.join(SCRIPT_DIR, "assets")
DB_PATH = os.path.join(SCRIPT_DIR, "registros_prensa_demo.db")

# --- CONFIGURACIÓN DE RED Y PLC ---
CONFIG_FILE = os.path.join(SCRIPT_DIR, "config_red.json")

def cargar_config_red():
    try:
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, "r") as f:
                return json.load(f)
    except Exception as e:
        print(f"Error al cargar config de red: {e}")
    return {"ip": "192.168.0.3 (SIMULADO)", "rack": 0, "slot": 1}

def guardar_config_red(ip, rack, slot):
    try:
        with open(CONFIG_FILE, "w") as f:
            json.dump({"ip": ip, "rack": int(rack), "slot": int(slot)}, f)
    except Exception as e:
        print(f"Error al guardar config de red: {e}")

OFFSET = 1960 	

# --- MAPEO COMPLETO DE MEMORIA ---
MAPEO = {
    'b002_on':      {'db': 1, 'start': 0,  'size': 2}, # VW0 (Fuerza Mínima)
    'b002_off':     {'db': 1, 'start': 2,  'size': 2}, # VW2 (Oculto)
    'b004_retardo': {'db': 1, 'start': 4,  'size': 2}, # VW4 (Segundo)
    'b001_ax':      {'db': 1, 'start': 6,  'size': 2}, # VW6 (Gráfica)
    'piston':       {'db': 1, 'start': 10, 'size': 1, 'bit': 0}, # V10.0
    
    # --- INDICADORES (SALIDAS DE RED EN BYTE 12) ---
    'barrera':      {'db': 1, 'start': 12, 'size': 1, 'bit': 0}, # V12.0
    'emergencia':   {'db': 1, 'start': 12, 'size': 1, 'bit': 1}, # V12.1
    'inicio_btn':   {'db': 1, 'start': 12, 'size': 1, 'bit': 2}  # V12.2
}

# --- PALETA DE COLORES PREMIUM (WKK) ---
COLOR_VERDE_WKK = "#0E8A3E"
COLOR_VERDE_OSCURO = "#0A6B2F"
COLOR_VERDE_CLARO = "#E2F0D9"
COLOR_FONDO = "#F8FAFC"      # Slate-50
COLOR_TARJETA = "#FFFFFF"    # White
COLOR_BORDE = "#E2E8F0"      # Slate-200
COLOR_TEXTO = "#1E293B"      # Slate-800
COLOR_TEXTO_SEC = "#64748B"  # Slate-500
COLOR_OK = "#10B981"         # Emerald-500 (Verde)
COLOR_NOK = "#EF4444"        # Red-500 (Rojo)


# --- MOCK PARA UTILIDADES DE SNAP7 ---
# Implementación nativa pura de python para evitar requerir el dll de snap7 en Windows
def get_int(buffer, byte_index):
    return int.from_bytes(buffer[byte_index:byte_index+2], byteorder='big', signed=True)

def set_int(buffer, byte_index, value):
    val_bytes = int(value).to_bytes(2, byteorder='big', signed=True)
    buffer[byte_index] = val_bytes[0]
    buffer[byte_index+1] = val_bytes[1]

def get_bool(buffer, byte_index, bit_index):
    return (buffer[byte_index] & (1 << bit_index)) != 0

def set_bool(buffer, byte_index, bit_index, value):
    if value:
        buffer[byte_index] |= (1 << bit_index)
    else:
        buffer[byte_index] &= ~(1 << bit_index)

class LogoClientMock:
    def __init__(self):
        self.connected = False
        # Valores simulados de registros en el PLC
        self.fuerza_minima = int(round(60 * 3.5)) + OFFSET  # Fuerza Mínima por defecto (60 kg, o 210 raw)
        self.fuerza_minima_off = int(round(56 * 3.5)) + OFFSET  # VW2
        self.retardo_centesimas = 250          # 2.50 segundos
        self.piston_active = False
        self.barrera_active = False
        self.emergencia_active = False
        self.inicio_active = False
        
        # Simulación de fuerza dinámica
        self.sim_time_in_cycle = 0.0

    def connect(self, ip, rack, slot):
        time.sleep(0.5)  # Simular latencia de red
        self.connected = True
        return True

    def get_connected(self):
        return self.connected

    def disconnect(self):
        self.connected = False

    def db_read(self, db_number, start_address, size_bytes):
        if not self.connected:
            raise Exception("PLC Connection Error")

        # vw0 (Fuerza Minima)
        if start_address == 0:
            buf = bytearray(2)
            set_int(buf, 0, self.fuerza_minima)
            return buf
        # vw2 (Oculto)
        elif start_address == 2:
            buf = bytearray(2)
            set_int(buf, 0, self.fuerza_minima_off)
            return buf
        # vw4 (Retardo)
        elif start_address == 4:
            buf = bytearray(2)
            set_int(buf, 0, self.retardo_centesimas)
            return buf
        # vw6 (Fuerza actual amplificada para grafica)
        elif start_address == 6:
            buf = bytearray(2)
            if self.piston_active:
                self.sim_time_in_cycle += 0.1
                # Simular curva de fuerza que sube
                # Sube por encima de la fuerza mínima para dar un ciclo OK
                target_peak = (self.fuerza_minima - OFFSET) + 18
                # Curva logarítmica de subida con ruido
                fuerza_calc = target_peak * (1.0 - math.exp(-self.sim_time_in_cycle * 2.0)) + random.uniform(-2, 2)
                if self.sim_time_in_cycle > 3.0: # Simular que se estabiliza
                    fuerza_calc = target_peak + random.uniform(-1, 1)
            else:
                self.sim_time_in_cycle = 0.0
                fuerza_calc = random.uniform(0.0, 5.0) # Ruido mínimo cerca de 0
                
            fuerza_calc = max(0, fuerza_calc)
            val_final = int(fuerza_calc) + OFFSET
            set_int(buf, 0, val_final)
            return buf
        # piston byte 10
        elif start_address == 10:
            buf = bytearray(1)
            set_bool(buf, 0, 0, self.piston_active)
            return buf
        # sensores byte 12
        elif start_address == 12:
            buf = bytearray(1)
            set_bool(buf, 0, 0, self.barrera_active)
            set_bool(buf, 0, 1, self.emergencia_active)
            set_bool(buf, 0, 2, self.inicio_active)
            return buf
        
        return bytearray(size_bytes)

    def db_write(self, db_number, start_address, buffer):
        if not self.connected:
            raise Exception("PLC Connection Error")

        # Escritura fuerza mínima
        if start_address == 0:
            self.fuerza_minima = get_int(buffer, 0)
        # Escritura VW2
        elif start_address == 2:
            self.fuerza_minima_off = get_int(buffer, 0)
        # Escritura retardo
        elif start_address == 4:
            self.retardo_centesimas = get_int(buffer, 0)
        # Escritura pistón (byte 10)
        elif start_address == 10:
            self.piston_active = get_bool(buffer, 0, 0)
        return True


class LogoHMI:
    def __init__(self, root):
        self.root = root
        self.root.title("WKK - HMI Sistema de Control y Monitoreo SCADA (DEMO WINDOWS)")
        
        # En modo demo en Windows, abrimos en una ventana grande pero no forzosamente fullscreen,
        # aunque incluimos el botón y soporte fullscreen si el usuario lo desea.
        self.root.geometry("1100x720")
        self.root.configure(bg=COLOR_FONDO)
        self.root.bind("<Escape>", lambda e: self.toggle_fullscreen())
        
        # Desenfocar entradas al hacer click en el fondo para ocultar el teclado
        self.root.bind("<Button-1>", self.click_afuera)

        # Instanciar el Cliente de PLC Simulado
        config_red = cargar_config_red()
        self.plc_ip = config_red["ip"]
        self.plc_rack = config_red["rack"]
        self.plc_slot = config_red["slot"]

        self.plc_client = LogoClientMock()
        self.is_connected = False
        self.running_loop = True  
        self.piston_state = False
        self.tick_counter = 0

        # Historial de gráfica
        self.grafica_datos = collections.deque([0]*50, maxlen=50)
        self.datos_limite = collections.deque([0]*50, maxlen=50)
        self.v6_filtrado_prev = 0

        # Estado del usuario
        self.perfil_actual = "operador"
        self.perfil_rol = "Operador"
        self.cycle_start_time = None
        
        # Historial y contadores de calidad
        self.total_ok_count = 0
        self.total_nok_count = 0
        self.last_piece_result = "--"
        self.last_piece_force = 0
        self.piston_last_state = False
        self.max_force_in_cycle = 0
        self.cycle_forces_list = []

        # Inicializar base de datos demo
        self.inicializar_db()

        # Construir Interfaz Gráfica
        self.create_widgets()
        
        # Aplicar permisos iniciales de Operador
        self.aplicar_permisos()

        # Iniciar la simulación en segundo plano
        threading.Thread(target=self.communication_loop, daemon=True).start()

        # Mostrar pantalla de Login al arrancar
        self.root.after(100, lambda: self.mostrar_login())

        # Vincular cambio de pestañas para refrescar tablas automáticamente
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)

    def establecer_pantalla_completa(self):
        try:
            # En Linux/X11, usar type toolbar elimina bordes y mantiene foco/entrada activa
            self.root.attributes('-type', 'toolbar')
        except Exception:
            try:
                self.root.overrideredirect(True)
            except Exception:
                pass
        
        try:
            screen_w = self.root.winfo_screenwidth()
            screen_h = self.root.winfo_screenheight()
            self.root.geometry(f"{screen_w}x{screen_h}+0+0")
            self.root.update_idletasks()
            self.root.focus_force()
        except Exception as e:
            print(f"Error al establecer pantalla completa: {e}")

    def toggle_fullscreen(self, event=None):
        try:
            is_borderless = self.root.overrideredirect()
            if is_borderless:
                self.root.overrideredirect(False)
                self.root.geometry("1100x720")
            else:
                self.establecer_pantalla_completa()
        except Exception:
            pass

    # --- BASE DE DATOS SQLite ---
    def inicializar_db(self):
        try:
            # Si el archivo de base de datos existe, verificar estructura
            if os.path.exists(DB_PATH):
                try:
                    conn = sqlite3.connect(DB_PATH)
                    cursor = conn.cursor()
                    cursor.execute("PRAGMA table_info(registros)")
                    cols = [c[1] for c in cursor.fetchall()]
                    conn.close()
                    
                    if cols and "resultado" not in cols:
                        # Estructura antigua detectada. Borrar archivo de base de datos para iniciar limpia.
                        os.remove(DB_PATH)
                        print("Base de datos antigua eliminada para actualizar estructura.")
                except Exception as e:
                    print(f"Error verificando estructura de DB: {e}")

            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS registros (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    fecha TEXT NOT NULL,
                    hora TEXT NOT NULL,
                    valor_fuerza INTEGER NOT NULL,
                    resultado TEXT NOT NULL,
                    limite_minimo INTEGER NOT NULL,
                    usuario TEXT NOT NULL
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS usuarios (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    password TEXT NOT NULL,
                    role TEXT NOT NULL
                )
            """)
            conn.commit()
            
            # Seed usuarios por defecto
            cursor.execute("SELECT COUNT(*) FROM usuarios")
            if cursor.fetchone()[0] == 0:
                cursor.execute("INSERT INTO usuarios (username, password, role) VALUES ('admin', 'admin', 'Administrador')")
                cursor.execute("INSERT INTO usuarios (username, password, role) VALUES ('operador', '', 'Operador')")
                conn.commit()
                
            conn.close()
            self.depurar_db_antigua()
        except Exception as e:
            print(f"Error al inicializar base de datos: {e}")

    def depurar_db_antigua(self):
        """Respalda y depura registros antiguos."""
        try:
            from datetime import timedelta
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            limite_fecha = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
            cursor.execute("SELECT COUNT(*) FROM registros WHERE fecha < ?", (limite_fecha,))
            count = cursor.fetchone()[0]
            
            if count > 0:
                backup_path = DB_PATH.replace(".db", "_backup.db")
                conn_backup = sqlite3.connect(backup_path)
                cursor_backup = conn_backup.cursor()
                cursor_backup.execute("""
                    CREATE TABLE IF NOT EXISTS registros (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        fecha TEXT NOT NULL,
                        hora TEXT NOT NULL,
                        valor_fuerza INTEGER NOT NULL,
                        resultado TEXT NOT NULL,
                        limite_minimo INTEGER NOT NULL,
                        usuario TEXT NOT NULL
                    )
                """)
                cursor.execute("SELECT fecha, hora, valor_fuerza, resultado, limite_minimo, usuario FROM registros WHERE fecha < ?", (limite_fecha,))
                for r in cursor.fetchall():
                    cursor_backup.execute("INSERT INTO registros (fecha, hora, valor_fuerza, resultado, limite_minimo, usuario) VALUES (?, ?, ?, ?, ?, ?)", r)
                conn_backup.commit()
                conn_backup.close()
                
                cursor.execute("DELETE FROM registros WHERE fecha < ?", (limite_fecha,))
                conn.commit()
                print(f"Respaldo automático demo: {count} registros antiguos depurados.")
            conn.close()
        except Exception as e:
            print(f"Error en depuración de DB: {e}")

    def registrar_ciclo_db(self, max_force, resultado, limite_minimo):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            ahora = datetime.now()
            cursor.execute("""
                INSERT INTO registros (fecha, hora, valor_fuerza, resultado, limite_minimo, usuario)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                ahora.strftime("%Y-%m-%d"),
                ahora.strftime("%H:%M:%S"),
                int(max_force),
                resultado,
                int(limite_minimo),
                self.perfil_actual
            ))
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"Error al registrar ciclo demo: {e}")

    # --- PANTALLAS Y LOGICA DE LOGIN ---
    def aplicar_permisos(self):
        if self.perfil_rol == "Operador":
            # Deshabilitar parámetros
            self.entry_v0.config(state="disabled")
            self.entry_v4.config(state="disabled")
            self.btn_write_v0.config(state="disabled")
            self.btn_write_v4.config(state="disabled")
            
            # Deshabilitar red
            self.entry_ip.config(state="disabled")
            self.entry_rack.config(state="disabled")
            self.entry_slot.config(state="disabled")
            self.btn_save_red.config(state="disabled")
            
            # Ocultar pestañas del notebook
            try: self.notebook.hide(self.tab_parametros)
            except Exception: pass
            try: self.notebook.hide(self.tab_registros)
            except Exception: pass
            try: self.notebook.hide(self.tab_usuarios)
            except Exception: pass
            
            # Mostrar vista simplificada para operador en monitoreo
            self.frame_monitoreo_admin.pack_forget()
            self.frame_monitoreo_operador.pack(fill="both", expand=True)
            
            self.notebook.select(self.tab_monitoreo)
            self.lbl_usuario_status.config(text=f"Perfil: Operador ({self.perfil_actual})", fg=COLOR_TEXTO_SEC)
        else:
            # Habilitar parámetros
            self.entry_v0.config(state="normal")
            self.entry_v4.config(state="normal")
            self.btn_write_v0.config(state="normal")
            self.btn_write_v4.config(state="normal")
            
            # Habilitar red
            self.entry_ip.config(state="normal")
            self.entry_rack.config(state="normal")
            self.entry_slot.config(state="normal")
            self.btn_save_red.config(state="normal")
            
            # Mostrar pestañas
            self.notebook.add(self.tab_parametros, text="Parámetros")
            self.notebook.add(self.tab_registros, text="Registros")
            self.notebook.add(self.tab_usuarios, text="Usuarios")
            
            # Mostrar vista completa para administrador en monitoreo
            self.frame_monitoreo_operador.pack_forget()
            self.frame_monitoreo_admin.pack(fill="both", expand=True)
            
            # Refrescar usuarios en admin
            self.refrescar_usuarios_gui()
            
            self.lbl_usuario_status.config(text=f"Perfil: Admin ({self.perfil_actual})", fg=COLOR_VERDE_WKK)

    def mostrar_login(self):
        # Frame de Login (Overlay que cubre toda la pantalla)
        self.login_overlay = tk.Frame(self.root, bg=COLOR_FONDO)
        self.login_overlay.place(relx=0, rely=0, relwidth=1, relheight=1)

        # Botón Cerrar App (para cerrar el programa desde el Login)
        btn_salir_login = tk.Button(self.login_overlay, text="✕", font=("Helvetica", 16, "bold"),
                                    fg=COLOR_TEXTO_SEC, bg=COLOR_FONDO, bd=0, padx=15, pady=8,
                                    activebackground=COLOR_NOK, activeforeground="white", command=self.on_closing)
        btn_salir_login.place(relx=1.0, rely=0.0, anchor="ne", x=-15, y=15)
        btn_salir_login.bind("<Enter>", lambda e: btn_salir_login.config(bg=COLOR_NOK, fg="white"))
        btn_salir_login.bind("<Leave>", lambda e: btn_salir_login.config(bg=COLOR_FONDO, fg=COLOR_TEXTO_SEC))

        card = tk.Frame(self.login_overlay, bg=COLOR_TARJETA, highlightbackground=COLOR_BORDE,
                        highlightthickness=1, padx=40, pady=40)
        card.place(relx=0.5, rely=0.5, anchor="center")

        # Cargar Logo
        try:
            if PIL_DISPONIBLE:
                wkk_img_login = Image.open(os.path.join(ASSETS_DIR, "Logo WKK.png"))
                wkk_ratio = wkk_img_login.width / wkk_img_login.height
                wkk_h = 60
                wkk_w = int(wkk_h * wkk_ratio)
                wkk_img_login = wkk_img_login.resize((wkk_w, wkk_h), RESAMPLE_METHOD)
                self.wkk_photo_login = ImageTk.PhotoImage(wkk_img_login)
                wkk_label_login = tk.Label(card, image=self.wkk_photo_login, bg=COLOR_TARJETA)
                wkk_label_login.pack(pady=(0, 15))
            else:
                raise ImportError
        except Exception:
            tk.Label(card, text="WKK", font=("Helvetica", 32, "bold"), fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(pady=(0, 10))

        tk.Label(card, text="SISTEMA DE CONTROL DE PRENSA", font=("Helvetica", 13, "bold"), fg=COLOR_TEXTO, bg=COLOR_TARJETA).pack(pady=(0, 4))
        tk.Label(card, text="Control de Acceso de Personal (DEMO)", font=("Helvetica", 10), fg=COLOR_TEXTO_SEC, bg=COLOR_TARJETA).pack(pady=(0, 20))

        # Perfil
        tk.Label(card, text="Perfil de Usuario:", font=("Helvetica", 10, "bold"), fg=COLOR_TEXTO, bg=COLOR_TARJETA).pack(anchor="w", pady=(0, 6))

        frame_perfiles = tk.Frame(card, bg=COLOR_TARJETA)
        frame_perfiles.pack(pady=(0, 15), fill="x")

        self.var_perfil = tk.StringVar(value="Operador")

        # Contenedor de campos de Administrador
        frame_admin_fields = tk.Frame(card, bg=COLOR_TARJETA)

        # Usuario
        lbl_user = tk.Label(frame_admin_fields, text="Usuario:", font=("Helvetica", 10, "bold"), fg=COLOR_TEXTO, bg=COLOR_TARJETA)
        lbl_user.pack(anchor="w", pady=(0, 4))
        entry_user = tk.Entry(frame_admin_fields, font=("Helvetica", 12), justify="center", width=25, relief="flat", highlightbackground=COLOR_BORDE, highlightthickness=1)
        entry_user.pack(pady=(0, 10))

        # Contraseña
        lbl_pass = tk.Label(frame_admin_fields, text="Contraseña:", font=("Helvetica", 10, "bold"), fg=COLOR_TEXTO, bg=COLOR_TARJETA)
        lbl_pass.pack(anchor="w", pady=(0, 4))
        entry_pass = tk.Entry(frame_admin_fields, show="*", font=("Helvetica", 12), justify="center", width=25, relief="flat", highlightbackground=COLOR_BORDE, highlightthickness=1)
        entry_pass.pack(pady=(0, 15))

        entry_user.bind("<Button-1>", lambda e: self.abrir_teclado_sistema())
        entry_pass.bind("<Button-1>", lambda e: self.abrir_teclado_sistema())

        def actualizar_estado_pass(perf):
            if perf == "Operador":
                entry_user.delete(0, tk.END)
                entry_pass.delete(0, tk.END)
                frame_admin_fields.pack_forget()
            else:
                frame_admin_fields.pack(after=frame_perfiles, fill="x", pady=(0, 10))
                entry_user.delete(0, tk.END)
                entry_user.insert(0, "admin")
                entry_pass.delete(0, tk.END)
                entry_user.focus_set()

        def seleccionar_perfil(perfil):
            self.var_perfil.set(perfil)
            if perfil == "Operador":
                btn_oper.config(bg=COLOR_VERDE_WKK, fg="white")
                btn_admin.config(bg="#F1F5F9", fg=COLOR_TEXTO_SEC)
                actualizar_estado_pass("Operador")
            else:
                btn_oper.config(bg="#F1F5F9", fg=COLOR_TEXTO_SEC)
                btn_admin.config(bg=COLOR_VERDE_WKK, fg="white")
                actualizar_estado_pass("Administrador")

        btn_oper = tk.Button(frame_perfiles, text="Operador", font=("Helvetica", 10, "bold"),
                             bd=0, pady=10, cursor="hand2", bg="#F1F5F9", fg=COLOR_TEXTO_SEC,
                             activebackground=COLOR_VERDE_OSCURO, activeforeground="white",
                             command=lambda: seleccionar_perfil("Operador"))
        btn_oper.pack(side="left", expand=True, fill="x", padx=(0, 4))

        btn_admin = tk.Button(frame_perfiles, text="Administrador", font=("Helvetica", 10, "bold"),
                              bd=0, pady=10, cursor="hand2", bg="#F1F5F9", fg=COLOR_TEXTO_SEC,
                              activebackground=COLOR_VERDE_OSCURO, activeforeground="white",
                              command=lambda: seleccionar_perfil("Administrador"))
        btn_admin.pack(side="left", expand=True, fill="x", padx=(4, 0))

        seleccionar_perfil("Operador")

        def intentar_login(event=None):
            perf = self.var_perfil.get()

            if perf == "Operador":
                self.perfil_actual = "operador"
                self.perfil_rol = "Operador"
                self.aplicar_permisos()
                self.login_overlay.destroy()
            elif perf == "Administrador":
                user_val = entry_user.get().strip()
                pass_val = entry_pass.get()
                
                try:
                    conn = sqlite3.connect(DB_PATH)
                    cursor = conn.cursor()
                    cursor.execute("SELECT password, role FROM usuarios WHERE username = ?", (user_val,))
                    row = cursor.fetchone()
                    conn.close()
                    
                    if row is not None:
                        db_pass, db_role = row
                        if db_pass == pass_val and db_role == "Administrador":
                            self.perfil_actual = user_val
                            self.perfil_rol = "Administrador"
                            self.aplicar_permisos()
                            self.login_overlay.destroy()
                        else:
                            messagebox.showerror("Error de Acceso", "Contraseña incorrecta o el usuario no es Administrador.", parent=self.login_overlay)
                    else:
                        messagebox.showerror("Error de Acceso", "Usuario no encontrado.", parent=self.login_overlay)
                except Exception as e:
                    messagebox.showerror("Error", f"Error al validar usuario:\n{e}", parent=self.login_overlay)

        btn_entrar = tk.Button(card, text="INICIAR SESIÓN", font=("Helvetica", 10, "bold"), fg="white", bg=COLOR_VERDE_WKK,
                               activebackground=COLOR_VERDE_OSCURO, activeforeground="white", bd=0, width=22, pady=10, command=intentar_login)
        btn_entrar.pack(pady=5)

        entry_user.bind("<Return>", lambda e: [self.cerrar_teclado_sistema(), intentar_login()])
        entry_pass.bind("<Return>", lambda e: [self.cerrar_teclado_sistema(), intentar_login()])
        entry_user.bind("<FocusOut>", lambda e: self.cerrar_teclado_sistema())
        entry_pass.bind("<FocusOut>", lambda e: self.cerrar_teclado_sistema())

    # --- CREACIÓN DE INTERFAZ GRÁFICA DE OPERACIÓN ---
    def create_widgets(self):
        # Contenedor Principal (Dashboard)
        self.dashboard = tk.Frame(self.root, bg=COLOR_FONDO)
        self.dashboard.pack(fill="both", expand=True)

        # 1. HEADER
        header = tk.Frame(self.dashboard, bg=COLOR_TARJETA, height=75)
        header.pack(fill="x", side="top")
        header.pack_propagate(False)

        # Logo WKK
        try:
            if PIL_DISPONIBLE:
                wkk_img = Image.open(os.path.join(ASSETS_DIR, "Logo WKK.png"))
                wkk_ratio = wkk_img.width / wkk_img.height
                wkk_h = 50
                wkk_w = int(wkk_h * wkk_ratio)
                wkk_img = wkk_img.resize((wkk_w, wkk_h), RESAMPLE_METHOD)
                self.wkk_photo = ImageTk.PhotoImage(wkk_img)
                wkk_label = tk.Label(header, image=self.wkk_photo, bg=COLOR_TARJETA)
                wkk_label.pack(side="left", padx=(15, 10), pady=12)
            else:
                raise ImportError
        except Exception:
            tk.Label(header, text="WKK", font=("Helvetica", 22, "bold"), fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(side="left", padx=15)

        # Separador vertical
        tk.Frame(header, bg=COLOR_VERDE_WKK, width=2, height=45).pack(side="left", padx=(5, 15), pady=15)

        tk.Label(header, text="HMI SCADA - Control de Prensa PLC (DEMO)", font=("Helvetica", 16, "bold"), fg=COLOR_TEXTO, bg=COLOR_TARJETA).pack(side="left", pady=15)

        # Botón Cerrar App (Extremo Derecho)
        btn_salir = tk.Button(header, text="✕", font=("Helvetica", 14, "bold"),
                              fg=COLOR_TEXTO_SEC, bg=COLOR_TARJETA, bd=0, padx=15, pady=8,
                              activebackground=COLOR_NOK, activeforeground="white", command=self.on_closing)
        btn_salir.pack(side="right", padx=15)
        btn_salir.bind("<Enter>", lambda e: btn_salir.config(bg=COLOR_NOK, fg="white"))
        btn_salir.bind("<Leave>", lambda e: btn_salir.config(bg=COLOR_TARJETA, fg=COLOR_TEXTO_SEC))

        # Información del Perfil
        frame_usuario = tk.Frame(header, bg=COLOR_TARJETA)
        frame_usuario.pack(side="right", padx=10, pady=5)

        self.lbl_usuario_status = tk.Label(frame_usuario, text="Perfil: Operador", font=("Helvetica", 11, "bold"), fg=COLOR_TEXTO_SEC, bg=COLOR_TARJETA)
        self.lbl_usuario_status.pack(side="left", padx=8)

        btn_switch_user = tk.Button(frame_usuario, text="Cambiar Usuario", font=("Helvetica", 11, "bold"), fg="white", bg=COLOR_VERDE_WKK, bd=0, padx=15, pady=8, command=self.mostrar_login)
        btn_switch_user.pack(side="left", padx=8)

        # Línea de acento verde
        tk.Frame(self.dashboard, bg=COLOR_VERDE_WKK, height=3).pack(fill="x", side="top")

        # 2. FOOTER
        footer = tk.Frame(self.dashboard, bg=COLOR_TARJETA, height=40)
        footer.pack(fill="x", side="bottom")
        footer.pack_propagate(False)

        # Línea de acento sobre el footer
        tk.Frame(self.dashboard, bg=COLOR_VERDE_WKK, height=1).pack(fill="x", side="bottom")

        # Logo Baluarte
        try:
            if PIL_DISPONIBLE:
                bal_img = Image.open(os.path.join(ASSETS_DIR, "Logo Horizontal sin fondo.png"))
                bal_ratio = bal_img.width / bal_img.height
                bal_h = 24
                bal_w = int(bal_h * bal_ratio)
                bal_img = bal_img.resize((bal_w, bal_h), RESAMPLE_METHOD)
                self.bal_photo = ImageTk.PhotoImage(bal_img)
                tk.Label(footer, image=self.bal_photo, bg=COLOR_TARJETA).pack(side="left", padx=10, pady=8)
            else:
                raise ImportError
        except Exception:
            tk.Label(footer, text="Desarrollado por Baluarte", font=("Helvetica", 8, "italic"), fg=COLOR_TEXTO_SEC, bg=COLOR_TARJETA).pack(side="left", padx=10)

        # Barra de Estado (Status Bar)
        self.status_bar = tk.Label(footer, text="Iniciando sistema...", font=("Helvetica", 10, "italic"), fg=COLOR_TEXTO_SEC, bg=COLOR_TARJETA, anchor='w')
        self.status_bar.pack(side="right", padx=15, pady=8)

        # 3. ÁREA DE CONTENIDO CENTRAL
        content = tk.Frame(self.dashboard, bg=COLOR_FONDO)
        content.pack(fill="both", expand=True, padx=12, pady=10)

        # Obtener dimensiones de pantalla para escalar el layout cómodamente
        screen_w = self.root.winfo_screenwidth()
        panel_w = 750 if screen_w >= 1800 else 480

        # Panel Izquierdo (Pestañas / Notebook)
        left_panel = tk.Frame(content, bg=COLOR_FONDO, width=panel_w)
        left_panel.pack(side="left", fill="y", padx=(0, 8))
        left_panel.pack_propagate(False)

        # Panel Derecho (Gráfica e Historial)
        right_panel = tk.Frame(content, bg=COLOR_FONDO)
        right_panel.pack(side="right", fill="both", expand=True)

        # Configuración del Notebook de pestañas
        style_nb = ttk.Style()
        style_nb.theme_use("clam")
        style_nb.configure("TNotebook", background=COLOR_FONDO, borderwidth=0)
        style_nb.configure("TNotebook.Tab", 
                           font=("Helvetica", 10, "bold"), 
                           padding=[12, 5], 
                           background="#E2E8F0", 
                           foreground=COLOR_TEXTO_SEC,
                           focuscolor="",
                           borderwidth=0,
                           relief="flat")
        style_nb.map("TNotebook.Tab",
                     background=[("selected", COLOR_VERDE_WKK), ("active", "#CBD5E1")],
                     foreground=[("selected", "white"), ("active", COLOR_TEXTO)])

        self.notebook = ttk.Notebook(left_panel, style="TNotebook")
        self.notebook.pack(fill="both", expand=True)

        self.tab_monitoreo = tk.Frame(self.notebook, bg=COLOR_FONDO)
        self.tab_parametros = tk.Frame(self.notebook, bg=COLOR_FONDO)
        self.tab_registros = tk.Frame(self.notebook, bg=COLOR_FONDO)
        self.tab_usuarios = tk.Frame(self.notebook, bg=COLOR_FONDO)

        self.notebook.add(self.tab_monitoreo, text="Monitoreo")
        self.notebook.add(self.tab_parametros, text="Parámetros")
        self.notebook.add(self.tab_registros, text="Registros")

        # --- PESTAÑA 1: MONITOREO ---
        # Contenedores para diferenciar la vista de Admin y Operador
        self.frame_monitoreo_admin = tk.Frame(self.tab_monitoreo, bg=COLOR_FONDO)
        self.frame_monitoreo_operador = tk.Frame(self.tab_monitoreo, bg=COLOR_FONDO)

        # --- VISTA ADMINISTRADOR ---
        # Card de Indicadores de Sensores (Admin)
        card_ind_admin = self.crear_tarjeta(self.frame_monitoreo_admin)
        card_ind_admin.pack(fill="x", pady=(0, 8))
        
        tk.Label(card_ind_admin, text="ESTADO DE SENSORES Y ENTRADAS", font=("Helvetica", 11, "bold"), fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(anchor="w", pady=(0, 10))

        self.c_piston, self.led_piston = self.crear_led_indicador(card_ind_admin, "Estado del Pistón")
        self.c_barrera, self.led_barrera = self.crear_led_indicador(card_ind_admin, "Sensor de Barrera")
        self.c_emergencia, self.led_emergencia = self.crear_led_indicador(card_ind_admin, "Paro de Emergencia")
        self.c_inicio, self.led_inicio = self.crear_led_indicador(card_ind_admin, "Inicio de Ciclo")

        # Card de Control de Pistón (Admin)
        card_ctrl_admin = self.crear_tarjeta(self.frame_monitoreo_admin)
        card_ctrl_admin.pack(fill="x", pady=0)

        tk.Label(card_ctrl_admin, text="CONTROL MANUAL DE PISTÓN", font=("Helvetica", 11, "bold"), fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(anchor="w", pady=(0, 10))

        frame_botones_admin = tk.Frame(card_ctrl_admin, bg=COLOR_TARJETA)
        frame_botones_admin.pack(fill="x", pady=5)

        self.btn_prender = tk.Button(frame_botones_admin, text="▲ PRENDER PISTÓN", font=("Helvetica", 10, "bold"), fg="white", bg=COLOR_OK, bd=0, pady=8, cursor="hand2", command=lambda: self.set_piston(True))
        self.btn_prender.pack(side="left", expand=True, fill="x", padx=(0, 5))
        self.btn_prender.bind("<Enter>", lambda e: self.btn_prender.config(bg="#059669"))
        self.btn_prender.bind("<Leave>", lambda e: self.btn_prender.config(bg=COLOR_OK))

        self.btn_apagar = tk.Button(frame_botones_admin, text="▼ APAGAR PISTÓN", font=("Helvetica", 10, "bold"), fg="white", bg=COLOR_NOK, bd=0, pady=8, cursor="hand2", command=lambda: self.set_piston(False))
        self.btn_apagar.pack(side="right", expand=True, fill="x", padx=(5, 0))
        self.btn_apagar.bind("<Enter>", lambda e: self.btn_apagar.config(bg="#DC2626"))
        self.btn_apagar.bind("<Leave>", lambda e: self.btn_apagar.config(bg=COLOR_NOK))

        # Card de Control de Calidad de la Pieza (Admin)
        card_calidad_admin = self.crear_tarjeta(self.frame_monitoreo_admin)
        card_calidad_admin.pack(fill="x", pady=(8, 0))
        tk.Label(card_calidad_admin, text="CONTROL DE CALIDAD DE LA PIEZA", font=("Helvetica", 11, "bold"), fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(anchor="w", pady=(0, 6))
        
        self.lbl_last_piece_admin = tk.Label(card_calidad_admin, text="ÚLTIMA PIEZA: --", font=("Helvetica", 14, "bold"), fg=COLOR_TEXTO_SEC, bg="#E2E8F0", pady=8)
        self.lbl_last_piece_admin.pack(fill="x", pady=(0, 6))

        self.lbl_last_force_admin = tk.Label(card_calidad_admin, text="Fuerza Registrada: --", font=("Helvetica", 11, "bold"), fg=COLOR_TEXTO, bg=COLOR_TARJETA)
        # self.lbl_last_force_admin.pack(anchor="w", pady=2)

        frame_counters_admin = tk.Frame(card_calidad_admin, bg=COLOR_TARJETA)
        frame_counters_admin.pack(fill="x", pady=4)

        self.lbl_counter_ok_admin = tk.Label(frame_counters_admin, text="OK: 0", font=("Helvetica", 12, "bold"), fg="white", bg=COLOR_OK, pady=6)
        self.lbl_counter_ok_admin.pack(side="left", expand=True, fill="x", padx=(0, 4))

        self.lbl_counter_nok_admin = tk.Label(frame_counters_admin, text="NOK: 0", font=("Helvetica", 12, "bold"), fg="white", bg=COLOR_NOK, pady=6)
        self.lbl_counter_nok_admin.pack(side="left", expand=True, fill="x", padx=(4, 0))

        # Card de Progreso y Fuerza (Admin)
        card_prog_admin = self.crear_tarjeta(self.frame_monitoreo_admin)
        card_prog_admin.pack(fill="x", pady=(8, 0))
        tk.Label(card_prog_admin, text="DURACIÓN DE CICLO", font=("Helvetica", 11, "bold"), fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(anchor="w", pady=(0, 5))
        
        self.lbl_progress_status_admin = tk.Label(card_prog_admin, text="Ciclo Inactivo", font=("Helvetica", 10, "bold"), fg=COLOR_TEXTO_SEC, bg=COLOR_TARJETA)
        self.lbl_progress_status_admin.pack(pady=(2, 2))

        self.frame_progress_admin = tk.Frame(card_prog_admin, bg="#E2E8F0", height=15)
        self.frame_progress_admin.pack(fill="x", pady=(2, 4))
        self.frame_progress_admin.pack_propagate(False)

        self.bar_fill_admin = tk.Frame(self.frame_progress_admin, bg=COLOR_VERDE_WKK, height=15)
        self.bar_fill_admin.place(x=0, y=0, width=0, height=15)

        # Recuadro para Fuerza Registrada para Administrador (en verde llamativo y muy visible)
        frame_prom_box_admin = tk.Frame(card_prog_admin, bg=COLOR_VERDE_CLARO, highlightbackground=COLOR_VERDE_WKK, highlightthickness=1.5, bd=0)
        frame_prom_box_admin.pack(fill="x", pady=(8, 2))
        
        tk.Label(frame_prom_box_admin, text="FUERZA REGISTRADA", font=("Helvetica", 10, "bold"), fg=COLOR_VERDE_OSCURO, bg=COLOR_VERDE_CLARO).pack(pady=(6, 2))
        self.lbl_average_display_admin = tk.Label(frame_prom_box_admin, text="-- kg", font=("Helvetica", 24, "bold"), fg=COLOR_VERDE_OSCURO, bg=COLOR_VERDE_CLARO)
        self.lbl_average_display_admin.pack(pady=(0, 6))

        # --- VISTA OPERADOR ---
        # Card 1: Estado del Proceso
        card_proc_oper = self.crear_tarjeta(self.frame_monitoreo_operador)
        card_proc_oper.pack(fill="x", pady=(0, 8))
        tk.Label(card_proc_oper, text="ESTADO DEL PROCESO", font=("Helvetica", 11, "bold"), fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(anchor="w", pady=(0, 8))
        self.lbl_piston_status_oper = tk.Label(card_proc_oper, text="PROCESO INACTIVO", font=("Helvetica", 14, "bold"), fg=COLOR_TEXTO_SEC, bg="#E2E8F0", pady=12)
        self.lbl_piston_status_oper.pack(fill="x")

        # Card 1.5: Control de Calidad de la Pieza
        card_calidad_oper = self.crear_tarjeta(self.frame_monitoreo_operador)
        card_calidad_oper.pack(fill="x", pady=(0, 8))
        tk.Label(card_calidad_oper, text="CONTROL DE CALIDAD DE LA PIEZA", font=("Helvetica", 11, "bold"), fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(anchor="w", pady=(0, 6))
        
        self.lbl_last_piece = tk.Label(card_calidad_oper, text="ÚLTIMA PIEZA: --", font=("Helvetica", 14, "bold"), fg=COLOR_TEXTO_SEC, bg="#E2E8F0", pady=8)
        self.lbl_last_piece.pack(fill="x", pady=(0, 6))

        self.lbl_last_force = tk.Label(card_calidad_oper, text="Fuerza Registrada: --", font=("Helvetica", 11, "bold"), fg=COLOR_TEXTO, bg=COLOR_TARJETA)
        # self.lbl_last_force.pack(anchor="w", pady=2)

        frame_counters = tk.Frame(card_calidad_oper, bg=COLOR_TARJETA)
        frame_counters.pack(fill="x", pady=4)

        self.lbl_counter_ok = tk.Label(frame_counters, text="OK: 0", font=("Helvetica", 12, "bold"), fg="white", bg=COLOR_OK, pady=6)
        self.lbl_counter_ok.pack(side="left", expand=True, fill="x", padx=(0, 4))

        self.lbl_counter_nok = tk.Label(frame_counters, text="NOK: 0", font=("Helvetica", 12, "bold"), fg="white", bg=COLOR_NOK, pady=6)
        self.lbl_counter_nok.pack(side="left", expand=True, fill="x", padx=(4, 0))

        # Card 2: Paro de Emergencia (LED Gigante)
        card_em_oper = self.crear_tarjeta(self.frame_monitoreo_operador)
        card_em_oper.pack(fill="x", pady=(0, 8))
        tk.Label(card_em_oper, text="PARO DE EMERGENCIA", font=("Helvetica", 11, "bold"), fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(anchor="w", pady=(0, 8))
        
        frame_em_layout = tk.Frame(card_em_oper, bg=COLOR_TARJETA)
        frame_em_layout.pack(fill="x")
        
        self.c_emergencia_oper = tk.Canvas(frame_em_layout, width=70, height=70, highlightthickness=0, bg=COLOR_TARJETA)
        self.c_emergencia_oper.pack(side="left", padx=(10, 20))
        self.led_emergencia_oper = self.c_emergencia_oper.create_oval(5, 5, 65, 65, fill="gray", outline="black", width=3)
        
        self.lbl_emergencia_text_oper = tk.Label(frame_em_layout, text="SISTEMA OK", font=("Helvetica", 15, "bold"), fg=COLOR_TEXTO, bg=COLOR_TARJETA)
        self.lbl_emergencia_text_oper.pack(side="left")

        # Card 3: Progreso de Ciclo (Barra de Progreso)
        card_prog_oper = self.crear_tarjeta(self.frame_monitoreo_operador)
        card_prog_oper.pack(fill="x", pady=(0, 8))
        tk.Label(card_prog_oper, text="DURACIÓN DE CICLO", font=("Helvetica", 11, "bold"), fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(anchor="w", pady=(0, 5))
        
        self.lbl_progress_status_oper = tk.Label(card_prog_oper, text="Ciclo Inactivo", font=("Helvetica", 10, "bold"), fg=COLOR_TEXTO_SEC, bg=COLOR_TARJETA)
        self.lbl_progress_status_oper.pack(pady=(2, 2))

        self.frame_progress_oper = tk.Frame(card_prog_oper, bg="#E2E8F0", height=15)
        self.frame_progress_oper.pack(fill="x", pady=(2, 4))
        self.frame_progress_oper.pack_propagate(False)

        self.bar_fill_oper = tk.Frame(self.frame_progress_oper, bg=COLOR_VERDE_WKK, height=15)
        self.bar_fill_oper.place(x=0, y=0, width=0, height=15)

        # Recuadro para Fuerza Registrada para Operador (en verde llamativo y muy visible)
        frame_prom_box_oper = tk.Frame(card_prog_oper, bg=COLOR_VERDE_CLARO, highlightbackground=COLOR_VERDE_WKK, highlightthickness=1.5, bd=0)
        frame_prom_box_oper.pack(fill="x", pady=(8, 2))
        
        tk.Label(frame_prom_box_oper, text="FUERZA REGISTRADA", font=("Helvetica", 10, "bold"), fg=COLOR_VERDE_OSCURO, bg=COLOR_VERDE_CLARO).pack(pady=(6, 2))
        self.lbl_average_display_oper = tk.Label(frame_prom_box_oper, text="-- kg", font=("Helvetica", 24, "bold"), fg=COLOR_VERDE_OSCURO, bg=COLOR_VERDE_CLARO)
        self.lbl_average_display_oper.pack(pady=(0, 6))



        # CARD EXTRA DE SIMULACIÓN DE ENTRADAS PLC (Sólo visible en la versión demo)
        card_sim = self.crear_tarjeta(self.tab_monitoreo)
        card_sim.pack(fill="x", pady=0)
        
        tk.Label(card_sim, text="⚙️ SIMULADOR DE HARDWARE PLC (DEMO)", font=("Helvetica", 10, "bold"), fg="#2196F3", bg=COLOR_TARJETA).pack(anchor="w", pady=(0, 8))
        
        frame_sim_checks = tk.Frame(card_sim, bg=COLOR_TARJETA)
        frame_sim_checks.pack(fill="x")

        self.var_sim_barrera = tk.BooleanVar(value=False)
        self.var_sim_emergencia = tk.BooleanVar(value=False)
        self.var_sim_inicio = tk.BooleanVar(value=False)

        def actualizar_simulacion():
            self.plc_client.barrera_active = self.var_sim_barrera.get()
            self.plc_client.emergencia_active = self.var_sim_emergencia.get()
            self.plc_client.inicio_active = self.var_sim_inicio.get()

        chk_barrera = tk.Checkbutton(frame_sim_checks, text="Activar Barrera", variable=self.var_sim_barrera, command=actualizar_simulacion, bg=COLOR_TARJETA, font=("Helvetica", 9), activebackground=COLOR_TARJETA)
        chk_barrera.pack(anchor="w", pady=2)

        chk_emergencia = tk.Checkbutton(frame_sim_checks, text="Activar Emergencia", variable=self.var_sim_emergencia, command=actualizar_simulacion, bg=COLOR_TARJETA, font=("Helvetica", 9), activebackground=COLOR_TARJETA)
        chk_emergencia.pack(anchor="w", pady=2)

        chk_inicio = tk.Checkbutton(frame_sim_checks, text="Pulsar Inicio de Ciclo", variable=self.var_sim_inicio, command=actualizar_simulacion, bg=COLOR_TARJETA, font=("Helvetica", 9), activebackground=COLOR_TARJETA)
        chk_inicio.pack(anchor="w", pady=2)

        # --- PESTAÑA 2: PARÁMETROS ---
        card_p = self.crear_tarjeta(self.tab_parametros)
        card_p.pack(fill="x", pady=0)

        tk.Label(card_p, text="AJUSTE DE PARÁMETROS PLC", font=("Helvetica", 11, "bold"), fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(anchor="w", pady=(0, 12))

        # Fuerza Mínima
        tk.Label(card_p, text="Fuerza Mínima:", font=("Helvetica", 10, "bold"), fg=COLOR_TEXTO, bg=COLOR_TARJETA).pack(anchor="w", pady=(2, 2))
        frame_v0 = tk.Frame(card_p, bg=COLOR_TARJETA)
        frame_v0.pack(fill="x", pady=(0, 10))

        self.lbl_v0 = tk.Label(frame_v0, text="--", font=("Helvetica", 14, "bold"), fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA, width=6, anchor="w")
        self.lbl_v0.pack(side="left")

        self.entry_v0 = tk.Entry(frame_v0, font=("Helvetica", 11), width=10, highlightbackground=COLOR_BORDE, highlightthickness=1, relief="flat", justify="center")
        self.entry_v0.pack(side="left", padx=10, ipady=3)

        self.btn_write_v0 = tk.Button(frame_v0, text="Guardar", font=("Helvetica", 10, "bold"), fg="white", bg=COLOR_VERDE_WKK, bd=0, padx=15, pady=8, command=lambda: self.escribir_vw('b002_on', self.entry_v0.get(), usar_offset=True))
        self.btn_write_v0.pack(side="left", padx=5)
        self.btn_write_v0.bind("<Enter>", lambda e: self.btn_write_v0.config(bg=COLOR_VERDE_OSCURO))
        self.btn_write_v0.bind("<Leave>", lambda e: self.btn_write_v0.config(bg=COLOR_VERDE_WKK))

        # Segundo
        tk.Label(card_p, text="Tiempo de Retardo:", font=("Helvetica", 10, "bold"), fg=COLOR_TEXTO, bg=COLOR_TARJETA).pack(anchor="w", pady=(2, 2))
        frame_v4 = tk.Frame(card_p, bg=COLOR_TARJETA)
        frame_v4.pack(fill="x", pady=0)

        self.lbl_v4 = tk.Label(frame_v4, text="-- s", font=("Helvetica", 14, "bold"), fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA, width=6, anchor="w")
        self.lbl_v4.pack(side="left")

        self.entry_v4 = tk.Entry(frame_v4, font=("Helvetica", 11), width=10, highlightbackground=COLOR_BORDE, highlightthickness=1, relief="flat", justify="center")
        self.entry_v4.pack(side="left", padx=10, ipady=3)

        self.btn_write_v4 = tk.Button(frame_v4, text="Guardar", font=("Helvetica", 10, "bold"), fg="white", bg=COLOR_VERDE_WKK, bd=0, padx=15, pady=8, command=lambda: self.escribir_retardo(self.entry_v4.get()))
        self.btn_write_v4.pack(side="left", padx=5)
        self.btn_write_v4.bind("<Enter>", lambda e: self.btn_write_v4.config(bg=COLOR_VERDE_OSCURO))
        self.btn_write_v4.bind("<Leave>", lambda e: self.btn_write_v4.config(bg=COLOR_VERDE_WKK))

        # Card 2: Configuración de Red (Solo Admin)
        card_red = self.crear_tarjeta(self.tab_parametros)
        card_red.pack(fill="x", pady=(10, 0))

        tk.Label(card_red, text="CONFIGURACIÓN DE RED Y PLC", font=("Helvetica", 11, "bold"), fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(anchor="w", pady=(0, 12))

        # IP del PLC
        frame_ip = tk.Frame(card_red, bg=COLOR_TARJETA)
        frame_ip.pack(fill="x", pady=5)
        tk.Label(frame_ip, text="IP del PLC:", font=("Helvetica", 10), fg=COLOR_TEXTO, bg=COLOR_TARJETA, width=15, anchor="w").pack(side="left")
        self.entry_ip = tk.Entry(frame_ip, font=("Helvetica", 10), width=18, highlightbackground=COLOR_BORDE, highlightthickness=1, relief="flat", justify="center")
        self.entry_ip.pack(side="left", padx=5)
        self.entry_ip.insert(0, self.plc_ip)

        # Rack / Slot
        frame_rack_slot = tk.Frame(card_red, bg=COLOR_TARJETA)
        frame_rack_slot.pack(fill="x", pady=5)
        
        tk.Label(frame_rack_slot, text="Rack:", font=("Helvetica", 10), fg=COLOR_TEXTO, bg=COLOR_TARJETA, width=6, anchor="w").pack(side="left")
        self.entry_rack = tk.Entry(frame_rack_slot, font=("Helvetica", 10), width=4, highlightbackground=COLOR_BORDE, highlightthickness=1, relief="flat", justify="center")
        self.entry_rack.pack(side="left", padx=5)
        self.entry_rack.insert(0, str(self.plc_rack))

        tk.Label(frame_rack_slot, text="Slot:", font=("Helvetica", 10), fg=COLOR_TEXTO, bg=COLOR_TARJETA, width=6, anchor="w").pack(side="left", padx=(15, 0))
        self.entry_slot = tk.Entry(frame_rack_slot, font=("Helvetica", 10), width=4, highlightbackground=COLOR_BORDE, highlightthickness=1, relief="flat", justify="center")
        self.entry_slot.pack(side="left", padx=5)
        self.entry_slot.insert(0, str(self.plc_slot))

        # Botón Guardar Red
        frame_btn_red = tk.Frame(card_red, bg=COLOR_TARJETA)
        frame_btn_red.pack(fill="x", pady=(10, 0))
        
        self.btn_save_red = tk.Button(frame_btn_red, text="Guardar y Reconectar", font=("Helvetica", 10, "bold"), fg="white", bg=COLOR_VERDE_WKK, bd=0, padx=15, pady=8, command=self.guardar_config_red_gui)
        self.btn_save_red.pack(side="right")

        # Botón para regresar a monitoreo
        frame_cerrar_p = tk.Frame(self.tab_parametros, bg=COLOR_FONDO)
        frame_cerrar_p.pack(fill="x", pady=(15, 0))
        btn_cerrar_p = tk.Button(frame_cerrar_p, text="✕ Cerrar Vista", font=("Helvetica", 11, "bold"), fg="white", bg="#64748B", bd=0, pady=10, padx=20, command=lambda: self.notebook.select(self.tab_monitoreo))
        btn_cerrar_p.pack(side="right")

        # --- PESTAÑA 3: REGISTROS ---
        card_filtros = self.crear_tarjeta(self.tab_registros)
        card_filtros.pack(fill="x", pady=(0, 6))
        
        tk.Label(card_filtros, text="FILTRAR HISTORIAL (DEMO)", font=("Helvetica", 10, "bold"), fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(anchor="w", pady=(0, 5))

        frame_grid = tk.Frame(card_filtros, bg=COLOR_TARJETA)
        frame_grid.pack(fill="x")

        # Fecha
        tk.Label(frame_grid, text="Fecha:", font=("Helvetica", 9), fg=COLOR_TEXTO, bg=COLOR_TARJETA).grid(row=0, column=0, sticky="w", pady=2, padx=(0, 6))
        self.entry_fecha_filtro = tk.Entry(frame_grid, font=("Helvetica", 9), width=10, highlightbackground=COLOR_BORDE, highlightthickness=1, relief="flat", justify="center")
        self.entry_fecha_filtro.grid(row=0, column=1, pady=2, sticky="w")
        self.entry_fecha_filtro.insert(0, datetime.now().strftime("%Y-%m-%d"))

        # Botón calendario modal
        btn_cal = tk.Button(frame_grid, text="📅", font=("Helvetica", 9, "bold"), fg="white", bg=COLOR_VERDE_WKK, bd=0, padx=8, pady=4, command=self.abrir_calendario_modal)
        btn_cal.grid(row=0, column=2, pady=2, padx=4)

        btn_hoy = tk.Button(frame_grid, text="Hoy", font=("Helvetica", 9, "bold"), fg=COLOR_TEXTO_SEC, bg="#F1F3F5", bd=0, padx=8, pady=4, command=self.set_fecha_hoy)
        btn_hoy.grid(row=0, column=3, pady=2)

        # Rango horas
        horas_list = [f"{h:02d}:00" for h in range(24)] + ["23:59"]
        tk.Label(frame_grid, text="Desde:", font=("Helvetica", 9), fg=COLOR_TEXTO, bg=COLOR_TARJETA).grid(row=1, column=0, sticky="w", pady=2)
        self.combo_h_inicio = ttk.Combobox(frame_grid, values=horas_list, width=6, font=("Helvetica", 9), state="readonly")
        self.combo_h_inicio.grid(row=1, column=1, pady=2, sticky="w")
        self.combo_h_inicio.set("00:00")

        tk.Label(frame_grid, text="Hasta:", font=("Helvetica", 9), fg=COLOR_TEXTO, bg=COLOR_TARJETA).grid(row=1, column=2, sticky="w", pady=2, padx=(4, 4))
        self.combo_h_fin = ttk.Combobox(frame_grid, values=horas_list, width=6, font=("Helvetica", 9), state="readonly")
        self.combo_h_fin.grid(row=1, column=3, pady=2, sticky="w")
        self.combo_h_fin.set("23:59")

        # Botonera filtros
        frame_filtros_btns = tk.Frame(card_filtros, bg=COLOR_TARJETA)
        frame_filtros_btns.pack(fill="x", pady=(5, 0))

        btn_filtrar = tk.Button(frame_filtros_btns, text="🔍 Filtrar", font=("Helvetica", 10, "bold"), fg="white", bg=COLOR_VERDE_WKK, bd=0, pady=8, command=self.aplicar_filtro_db)
        btn_filtrar.pack(side="left", expand=True, fill="x", padx=(0, 4))

        btn_limpiar = tk.Button(frame_filtros_btns, text="Limpiar", font=("Helvetica", 10, "bold"), fg=COLOR_TEXTO_SEC, bg="#E2E8F0", bd=0, pady=8, command=self.limpiar_filtro_db)
        btn_limpiar.pack(side="right", expand=True, fill="x", padx=(4, 0))

        # Card Tabla Registros
        card_tabla = tk.Frame(self.tab_registros, bg=COLOR_TARJETA, highlightbackground=COLOR_BORDE, highlightthickness=1, padx=8, pady=8)
        card_tabla.pack(fill="both", expand=True, pady=(6, 0))

        scroll_y = ttk.Scrollbar(card_tabla, orient="vertical")
        columns = ("id", "fecha", "hora", "fuerza", "resultado", "limite_minimo", "usuario")
        self.tree_registros = ttk.Treeview(card_tabla, columns=columns, show="headings", yscrollcommand=scroll_y.set, height=6)
        scroll_y.config(command=self.tree_registros.yview)

        headers = {
            "id": "ID", "fecha": "Fecha", "hora": "Hora", "fuerza": "Fuerza Máxima",
            "resultado": "Resultado", "limite_minimo": "Límite Mínimo (VW0)", "usuario": "Usuario"
        }
        for col, val in headers.items():
            self.tree_registros.heading(col, text=val)
            self.tree_registros.column(col, anchor="center", width=80)
        self.tree_registros.column("id", width=35)
        self.tree_registros.column("resultado", width=90)
        self.tree_registros.column("usuario", width=80)

        self.tree_registros.pack(side="left", fill="both", expand=True)
        scroll_y.pack(side="right", fill="y")

        # Botonera de acciones de registros
        frame_acc = tk.Frame(self.tab_registros, bg=COLOR_FONDO)
        frame_acc.pack(fill="x", pady=(6, 0))

        btn_export = tk.Button(frame_acc, text="📊 Exportar a Excel", font=("Helvetica", 11, "bold"), fg="white", bg=COLOR_VERDE_WKK, bd=0, pady=10, command=self.exportar_a_excel)
        btn_export.pack(side="left", expand=True, fill="x", padx=(0, 4))
        btn_export.bind("<Enter>", lambda e: btn_export.config(bg=COLOR_VERDE_OSCURO))
        btn_export.bind("<Leave>", lambda e: btn_export.config(bg=COLOR_VERDE_WKK))

        self.btn_eliminar = tk.Button(frame_acc, text="🗑 Limpiar Registros", font=("Helvetica", 11, "bold"), fg="white", bg=COLOR_NOK, bd=0, pady=10, command=self.confirmar_y_eliminar_db)
        self.btn_eliminar.pack(side="right", expand=True, fill="x", padx=(4, 0))
        self.btn_eliminar.bind("<Enter>", lambda e: self.btn_eliminar.config(bg="#DC2626"))
        self.btn_eliminar.bind("<Leave>", lambda e: self.btn_eliminar.config(bg=COLOR_NOK))

        # Botón para regresar a monitoreo
        frame_cerrar_reg = tk.Frame(self.tab_registros, bg=COLOR_FONDO)
        frame_cerrar_reg.pack(fill="x", pady=(15, 0))
        btn_cerrar_reg = tk.Button(frame_cerrar_reg, text="✕ Cerrar Vista", font=("Helvetica", 11, "bold"), fg="white", bg="#64748B", bd=0, pady=10, padx=20, command=lambda: self.notebook.select(self.tab_monitoreo))
        btn_cerrar_reg.pack(side="right")

        # --- PESTAÑA 4: GESTIÓN DE USUARIOS (Solo Admin) ---
        self.tab_usuarios = tk.Frame(self.notebook, bg=COLOR_FONDO)

        frame_us_left = tk.Frame(self.tab_usuarios, bg=COLOR_FONDO, width=220)
        frame_us_left.pack(side="left", fill="both", expand=True, padx=(0, 6))

        frame_us_right = tk.Frame(self.tab_usuarios, bg=COLOR_FONDO, width=220)
        frame_us_right.pack(side="right", fill="both", expand=True, padx=(6, 0))

        # Tarjeta Lista de Usuarios
        card_us_list = self.crear_tarjeta(frame_us_left)
        card_us_list.pack(fill="both", expand=True)
        tk.Label(card_us_list, text="USUARIOS DEL SISTEMA", font=("Helvetica", 10, "bold"), fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(anchor="w", pady=(0, 6))

        scroll_us = ttk.Scrollbar(card_us_list, orient="vertical")
        self.tree_usuarios = ttk.Treeview(card_us_list, columns=("usuario", "rol"), show="headings", yscrollcommand=scroll_us.set, height=10)
        scroll_us.config(command=self.tree_usuarios.yview)
        
        self.tree_usuarios.heading("usuario", text="Usuario")
        self.tree_usuarios.heading("rol", text="Rol")
        self.tree_usuarios.column("usuario", anchor="center", width=100)
        self.tree_usuarios.column("rol", anchor="center", width=100)
        
        self.tree_usuarios.pack(side="left", fill="both", expand=True)
        scroll_us.pack(side="right", fill="y")
        self.tree_usuarios.bind("<<TreeviewSelect>>", self.on_usuario_select)

        # Tarjeta 1: Crear Usuario
        card_us_create = self.crear_tarjeta(frame_us_right)
        card_us_create.pack(fill="x", pady=(0, 6))
        tk.Label(card_us_create, text="CREAR NUEVO USUARIO", font=("Helvetica", 10, "bold"), fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(anchor="w", pady=(0, 8))

        tk.Label(card_us_create, text="Nombre de Usuario:", font=("Helvetica", 9), fg=COLOR_TEXTO, bg=COLOR_TARJETA).pack(anchor="w")
        self.entry_new_user = tk.Entry(card_us_create, font=("Helvetica", 10), highlightbackground=COLOR_BORDE, highlightthickness=1, relief="flat")
        self.entry_new_user.pack(fill="x", pady=(2, 6), ipady=2)

        tk.Label(card_us_create, text="Contraseña:", font=("Helvetica", 9), fg=COLOR_TEXTO, bg=COLOR_TARJETA).pack(anchor="w")
        self.entry_new_pass = tk.Entry(card_us_create, font=("Helvetica", 10), show="*", highlightbackground=COLOR_BORDE, highlightthickness=1, relief="flat")
        self.entry_new_pass.pack(fill="x", pady=(2, 6), ipady=2)

        tk.Label(card_us_create, text="Rol del Usuario:", font=("Helvetica", 9), fg=COLOR_TEXTO, bg=COLOR_TARJETA).pack(anchor="w")
        self.combo_new_role = ttk.Combobox(card_us_create, values=["Administrador", "Operador"], state="readonly", font=("Helvetica", 9))
        self.combo_new_role.pack(fill="x", pady=(2, 8))
        self.combo_new_role.set("Administrador")

        btn_create_us = tk.Button(card_us_create, text="➕ Crear Usuario", font=("Helvetica", 10, "bold"), fg="white", bg=COLOR_VERDE_WKK, bd=0, pady=8, command=self.crear_usuario_gui)
        btn_create_us.pack(fill="x")

        # Tarjeta 2: Cambiar Contraseña / Eliminar Usuario
        card_us_edit = self.crear_tarjeta(frame_us_right)
        card_us_edit.pack(fill="x")
        tk.Label(card_us_edit, text="EDITAR / ELIMINAR SELECCIONADO", font=("Helvetica", 10, "bold"), fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(anchor="w", pady=(0, 8))

        self.lbl_selected_user_name = tk.Label(card_us_edit, text="Selecciona un usuario de la lista", font=("Helvetica", 9, "italic"), fg=COLOR_TEXTO_SEC, bg=COLOR_TARJETA)
        self.lbl_selected_user_name.pack(anchor="w", pady=(0, 6))

        tk.Label(card_us_edit, text="Nueva Contraseña:", font=("Helvetica", 9), fg=COLOR_TEXTO, bg=COLOR_TARJETA).pack(anchor="w")
        self.entry_edit_pass = tk.Entry(card_us_edit, font=("Helvetica", 10), show="*", highlightbackground=COLOR_BORDE, highlightthickness=1, relief="flat")
        self.entry_edit_pass.pack(fill="x", pady=(2, 8), ipady=2)

        frame_edit_btns = tk.Frame(card_us_edit, bg=COLOR_TARJETA)
        frame_edit_btns.pack(fill="x")

        self.btn_save_pass = tk.Button(frame_edit_btns, text="Guardar Pass", font=("Helvetica", 10, "bold"), fg="white", bg=COLOR_VERDE_WKK, bd=0, pady=8, command=self.guardar_contrasena_gui)
        self.btn_save_pass.pack(side="left", expand=True, fill="x", padx=(0, 3))

        self.btn_delete_us = tk.Button(frame_edit_btns, text="Eliminar", font=("Helvetica", 10, "bold"), fg="white", bg=COLOR_NOK, bd=0, pady=8, command=self.eliminar_usuario_gui)
        self.btn_delete_us.pack(side="right", expand=True, fill="x", padx=(3, 0))

        # Botón para regresar a monitoreo
        frame_cerrar_us = tk.Frame(frame_us_right, bg=COLOR_FONDO)
        frame_cerrar_us.pack(fill="x", pady=(15, 0))
        btn_cerrar_us = tk.Button(frame_cerrar_us, text="✕ Cerrar Vista", font=("Helvetica", 11, "bold"), fg="white", bg="#64748B", bd=0, pady=10, padx=20, command=lambda: self.notebook.select(self.tab_monitoreo))
        btn_cerrar_us.pack(side="right")

        # --- PANEL DERECHO: GRÁFICA E HISTORIAL ---
        card_grafica = self.crear_tarjeta(right_panel)
        card_grafica.pack(fill="both", expand=True, pady=(0, 6))

        tk.Label(card_grafica, text="B001 Amplificador (VW6) - Tiempo Real", font=("Helvetica", 11, "bold"), fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(anchor="w", pady=(0, 5))

        self.fig = Figure(figsize=(5, 3.2), dpi=100)
        self.fig.patch.set_facecolor(COLOR_TARJETA)
        self.ax = self.fig.add_subplot(111)
        self.ax.set_facecolor('#FAFAFA')
        self.ax.set_ylim(0, 100) 
        self.ax.grid(True, linestyle="--", alpha=0.3, color="#ADB5BD")
        
        # Eliminar bordes para estética limpia
        self.ax.spines['top'].set_visible(False)
        self.ax.spines['right'].set_visible(False)
        self.ax.spines['left'].set_color(COLOR_BORDE)
        self.ax.spines['bottom'].set_color(COLOR_BORDE)
        self.ax.tick_params(colors=COLOR_TEXTO_SEC, labelsize=9)

        self.line, = self.ax.plot(self.grafica_datos, color=COLOR_VERDE_WKK, linewidth=2.5, label="Fuerza")
        self.line_limite, = self.ax.plot(self.datos_limite, color="#2196F3", linestyle="--", linewidth=1.5, label="Fuerza Mínima")
        self.ax.legend(loc="upper left", fontsize=9, framealpha=0.9, edgecolor=COLOR_BORDE)
        self.fig.tight_layout(pad=1.5)

        self.canvas = FigureCanvasTkAgg(self.fig, master=card_grafica)
        self.canvas.get_tk_widget().pack(fill='both', expand=True)

        self.lbl_v6_actual = tk.Label(card_grafica, text="Valor actual: --", font=('Helvetica', 12, 'bold'), fg=COLOR_TEXTO, bg=COLOR_TARJETA)
        self.lbl_v6_actual.pack(pady=4)

        # Historial de lista
        card_log = self.crear_tarjeta(right_panel)
        card_log.pack(fill="x", pady=0)
        card_log.pack_propagate(False)
        card_log.config(height=140)

        tk.Label(card_log, text="Historial Numérico (1 muestra por seg.)", font=("Helvetica", 10, "bold"), fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(anchor="w", pady=(0, 4))
        
        frame_log = tk.Frame(card_log, bg=COLOR_TARJETA)
        frame_log.pack(fill='both', expand=True)
        
        self.scrollbar = ttk.Scrollbar(frame_log)
        self.scrollbar.pack(side='right', fill='y')
        
        self.listbox_log = tk.Listbox(frame_log, yscrollcommand=self.scrollbar.set, font=('Consolas', 9), relief="flat", highlightbackground=COLOR_BORDE, highlightthickness=1)
        self.listbox_log.pack(side='left', fill='both', expand=True)
        self.scrollbar.config(command=self.listbox_log.yview)

        # Vincular campos de entrada con el Teclado Virtual del Sistema (Raspberry Pi OS)
        for entry in [self.entry_v0, self.entry_v4, self.entry_ip, self.entry_rack, self.entry_slot, 
                      self.entry_fecha_filtro, self.entry_new_user, self.entry_new_pass, self.entry_edit_pass]:
            entry.bind("<Button-1>", lambda e: self.abrir_teclado_sistema())
            entry.bind("<Return>", lambda e: self.cerrar_teclado_sistema())
            entry.bind("<FocusOut>", lambda e: self.cerrar_teclado_sistema())

    def on_tab_changed(self, event):
        tab_id = self.notebook.select()
        if not tab_id:
            return
        tab_text = self.notebook.tab(tab_id, "text")
        if tab_text == "Registros":
            self.refrescar_tabla_gui()
        elif tab_text == "Usuarios":
            self.refrescar_usuarios_gui()

    def abrir_teclado_sistema(self, event=None):
        # Intentar mostrar el teclado virtual predeterminado de Raspberry Pi OS (Squeekboard) vía D-Bus
        try:
            subprocess.Popen([
                "gdbus", "call",
                "--session",
                "--dest", "sm.puri.OSK0",
                "--object-path", "/sm/puri/OSK0",
                "--method", "sm.puri.OSK0.SetVisible",
                "true"
            ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception:
            pass

        try:
            subprocess.Popen([
                "busctl", "call",
                "--user", "sm.puri.OSK0",
                "/sm/puri/OSK0",
                "sm.puri.OSK0", "SetVisible", "b", "true"
            ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception:
            pass

    def cerrar_teclado_sistema(self, event=None):
        # Intentar ocultar el teclado virtual predeterminado de Raspberry Pi OS (Squeekboard) vía D-Bus
        try:
            subprocess.Popen([
                "gdbus", "call",
                "--session",
                "--dest", "sm.puri.OSK0",
                "--object-path", "/sm/puri/OSK0",
                "--method", "sm.puri.OSK0.SetVisible",
                "false"
            ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception:
            pass

        try:
            subprocess.Popen([
                "busctl", "call",
                "--user", "sm.puri.OSK0",
                "/sm/puri/OSK0",
                "sm.puri.OSK0", "SetVisible", "b", "false"
            ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception:
            pass

    def click_afuera(self, event):
        # Si el click ocurre fuera de cualquier Entry, removemos el foco
        widget = event.widget
        if widget and not isinstance(widget, tk.Entry):
            self.root.focus_set()

    # --- UI HELPERS ---
    def crear_tarjeta(self, parent, **kwargs):
        return tk.Frame(parent, bg=COLOR_TARJETA, highlightbackground=COLOR_BORDE, highlightthickness=1, padx=15, pady=10, **kwargs)

    def crear_led_indicador(self, parent, texto):
        frame = ttk.Frame(parent, style='Card.TFrame')
        frame.pack(fill='x', pady=3)
        canvas = tk.Canvas(frame, width=28, height=28, highlightthickness=0)
        canvas.pack(side='left', padx=10)
        led = canvas.create_oval(4, 4, 24, 24, fill="gray", outline="black", width=2)
        ttk.Label(frame, text=texto, font=('Helvetica', 10)).pack(side='left', padx=5)
        return canvas, led

    # --- COMUNICACION CON PLC (HILO DE FONDO SIMULADO) ---
    def communication_loop(self):
        while self.running_loop:
            if not self.is_connected:
                self.root.after(0, self.update_status_gui, f"Intentando conectar a {self.plc_ip}...", "orange")
                # Simular tiempo de conexion
                time.sleep(1.5)
                self.plc_client.connect(self.plc_ip, self.plc_rack, self.plc_slot)
                if self.plc_client.get_connected():
                    self.is_connected = True
                    self.root.after(0, self.update_status_gui, f"Conectado a {self.plc_ip} - PLC Simulado OK", "green")
            
            if self.is_connected:
                self.read_cycle()
            else:
                self.root.after(0, self.reset_interface_to_gray)
                time.sleep(3)

    def read_cycle(self):
        try:
            raw_v0 = self.plc_client.db_read(MAPEO['b002_on']['db'], MAPEO['b002_on']['start'], MAPEO['b002_on']['size'])
            v0 = int(round(max(0, get_int(raw_v0, 0) - OFFSET) / 3.5))

            self.plc_client.db_read(MAPEO['b002_off']['db'], MAPEO['b002_off']['start'], MAPEO['b002_off']['size'])

            raw_v4 = self.plc_client.db_read(MAPEO['b004_retardo']['db'], MAPEO['b004_retardo']['start'], MAPEO['b004_retardo']['size'])
            v4_segundos = get_int(raw_v4, 0) / 100.0

            raw_v6 = self.plc_client.db_read(MAPEO['b001_ax']['db'], MAPEO['b001_ax']['start'], MAPEO['b001_ax']['size'])
            v6 = int(round(max(0, get_int(raw_v6, 0) - OFFSET) / 3.5))

            raw_piston = self.plc_client.db_read(MAPEO['piston']['db'], MAPEO['piston']['start'], MAPEO['piston']['size'])
            p_act = get_bool(raw_piston, 0, MAPEO['piston']['bit'])

            raw_byte12 = self.plc_client.db_read(MAPEO['barrera']['db'], MAPEO['barrera']['start'], 1)
            b_act = get_bool(raw_byte12, 0, MAPEO['barrera']['bit'])
            e_act = get_bool(raw_byte12, 0, MAPEO['emergencia']['bit'])
            i_act = get_bool(raw_byte12, 0, MAPEO['inicio_btn']['bit'])

            self.root.after(0, self.update_gui, v0, v4_segundos, v6, p_act, b_act, e_act, i_act)
            time.sleep(0.1)
        except Exception as e:
            print(f"Error de red detectado: {e}")
            self.is_connected = False
            self.plc_client.disconnect()

    def guardar_config_red_gui(self):
        ip = self.entry_ip.get().strip()
        rack_str = self.entry_rack.get().strip()
        slot_str = self.entry_slot.get().strip()

        if not ip:
            messagebox.showerror("Error", "La dirección IP no puede estar vacía.")
            return

        try:
            rack = int(rack_str)
            slot = int(slot_str)
        except ValueError:
            messagebox.showerror("Error", "Rack y Slot deben ser números enteros.")
            return

        guardar_config_red(ip, rack, slot)

        self.plc_ip = ip
        self.plc_rack = rack
        self.plc_slot = slot

        self.is_connected = False
        try:
            self.plc_client.disconnect()
        except Exception:
            pass

        messagebox.showinfo("Configuración de Red", f"Configuración guardada.\nIntentando conectar a {ip}...")

    def update_status_gui(self, texto, color):
        self.status_bar.config(text=texto, foreground=color)

    def reset_interface_to_gray(self):
        self.lbl_v0.config(text="--")
        self.lbl_v4.config(text="-- s")
        self.lbl_v6_actual.config(text="Valor actual: --")
        for c, l in [(self.c_piston, self.led_piston), (self.c_barrera, self.led_barrera), 
                     (self.c_emergencia, self.led_emergencia), (self.c_inicio, self.led_inicio)]:
            c.itemconfig(l, fill="gray", outline="black")

    def set_piston(self, estado_deseado):
        if not self.is_connected: return
        try:
            buffer = self.plc_client.db_read(MAPEO['piston']['db'], MAPEO['piston']['start'], MAPEO['piston']['size'])
            set_bool(buffer, 0, MAPEO['piston']['bit'], estado_deseado)
            self.plc_client.db_write(MAPEO['piston']['db'], MAPEO['piston']['start'], buffer)
        except Exception as e:
            print(f"Error escribiendo pistón simulado: {e}")

    def escribir_vw(self, vw_name, valor_str, usar_offset):
        if not self.is_connected:
            messagebox.showwarning("Aviso", "Sin conexión al PLC.")
            return
        try:
            valor_entero = int(valor_str)
            if vw_name == 'b002_on':
                valor_final = int(round(valor_entero * 3.5)) + OFFSET
            else:
                valor_final = valor_entero + OFFSET if usar_offset else valor_entero
                
            buffer = bytearray(2)
            set_int(buffer, 0, valor_final)
            self.plc_client.db_write(MAPEO[vw_name]['db'], MAPEO[vw_name]['start'], buffer)
            
            if vw_name == 'b002_on':
                valor_v2 = int(round(valor_entero * 3.5)) - 10
                valor_final_v2 = valor_v2 + OFFSET 
                buffer_v2 = bytearray(2)
                set_int(buffer_v2, 0, valor_final_v2)
                self.plc_client.db_write(MAPEO['b002_off']['db'], MAPEO['b002_off']['start'], buffer_v2)
                self.update_status_gui(f"Fuerza Mínima modificada a {valor_entero}.", "blue")
        except ValueError:
            messagebox.showerror("Error", "Ingresa un número entero válido.")
        except Exception as e:
            print(f"Fallo de escritura: {e}")

    def escribir_retardo(self, segundos_str):
        if not self.is_connected:
            messagebox.showwarning("Aviso", "Sin conexión al PLC.")
            return
        try:
            segundos = float(segundos_str)
            centesimas = int(segundos * 100)
            buffer = bytearray(2)
            set_int(buffer, 0, centesimas)
            self.plc_client.db_write(MAPEO['b004_retardo']['db'], MAPEO['b004_retardo']['start'], buffer)
            self.update_status_gui(f"Segundo (VW4) actualizado a {segundos} s.", "blue")
        except ValueError:
            messagebox.showerror("Error", "Ingresa un tiempo válido.")
        except Exception as e:
            print(f"Fallo de escritura: {e}")

    # --- ACTUALIZACIÓN DE INTERFAZ ---
    def update_gui(self, v0, v4, v6, p_act, b_act, e_act, i_act):
        if not self.is_connected: return

        # Aplicar filtro de promedio móvil exponencial (EMA) para suavizar la fuerza y remover picos rápidos
        alpha = 0.35  # Ajuste de suavizado
        v6_filtrado = int(round(alpha * v6 + (1 - alpha) * self.v6_filtrado_prev))
        self.v6_filtrado_prev = v6_filtrado

        self.lbl_v0.config(text=str(v0))
        self.lbl_v4.config(text=f"{v4:.1f} s")
        self.lbl_v6_actual.config(text=f"Valor actual: {v6_filtrado}")

        # --- DETECCIÓN DE CICLOS (Proceso activo si la fuerza supera los 15 kg o el pistón se activa) ---
        proceso_activo = (v6_filtrado > 15) or p_act
        
        if proceso_activo:
            if self.cycle_start_time is None:
                self.cycle_start_time = time.time()
                self.cycle_forces_list = []  # Inicializar lista vacía para ignorar el golpe inicial
                self.grafica_datos.clear()   # Limpiar trazado de gráfica anterior al iniciar ciclo
                self.datos_limite.clear()
                if hasattr(self, 'lbl_average_display_oper') and self.lbl_average_display_oper.winfo_exists():
                    self.lbl_average_display_oper.config(text="Prensando...")
                if hasattr(self, 'lbl_average_display_admin') and self.lbl_average_display_admin.winfo_exists():
                    self.lbl_average_display_admin.config(text="Prensando...")
            else:
                # Omitir el pico de impacto inicial ignorando lecturas de los primeros 0.4 segundos
                elapsed_time = time.time() - self.cycle_start_time
                if elapsed_time > 0.4:
                    self.cycle_forces_list.append(v6_filtrado)
        else:
            if self.cycle_start_time is not None:
                # El ciclo acaba de terminar. Evaluamos la duración total del ciclo
                elapsed_total = time.time() - self.cycle_start_time
                
                # Si duró al menos 0.4 segundos, es un ciclo de prensado real (no un golpe por colocación o ruido)
                if elapsed_total >= 0.4:
                    if self.cycle_forces_list:
                        average_force = max(self.cycle_forces_list)
                    else:
                        average_force = v6_filtrado
                    
                    resultado = "OK" if average_force >= v0 else "NOK"
                    
                    # Registrar en base de datos
                    self.registrar_ciclo_db(average_force, resultado, v0)
                    
                    # Actualizar contadores
                    if resultado == "OK":
                        self.total_ok_count += 1
                    else:
                        self.total_nok_count += 1
                        
                    self.last_piece_result = resultado
                    self.last_piece_force = average_force
                    
                    # Refrescar labels del operador
                    if hasattr(self, 'lbl_last_piece') and self.lbl_last_piece.winfo_exists():
                        self.lbl_last_piece.config(text=f"ÚLTIMA PIEZA: {resultado}", bg=COLOR_OK if resultado == "OK" else COLOR_NOK, fg="white")
                        self.lbl_last_force.config(text=f"Fuerza Registrada: {average_force} kg")
                        self.lbl_counter_ok.config(text=f"OK: {self.total_ok_count}")
                        self.lbl_counter_nok.config(text=f"NOK: {self.total_nok_count}")
                    
                    # Refrescar labels del administrador
                    if hasattr(self, 'lbl_last_piece_admin') and self.lbl_last_piece_admin.winfo_exists():
                        self.lbl_last_piece_admin.config(text=f"ÚLTIMA PIEZA: {resultado}", bg=COLOR_OK if resultado == "OK" else COLOR_NOK, fg="white")
                        self.lbl_last_force_admin.config(text=f"Fuerza Registrada: {average_force} kg")
                        self.lbl_counter_ok_admin.config(text=f"OK: {self.total_ok_count}")
                        self.lbl_counter_nok_admin.config(text=f"NOK: {self.total_nok_count}")

                    # Actualizar indicadores verdes grandes
                    if hasattr(self, 'lbl_average_display_oper') and self.lbl_average_display_oper.winfo_exists():
                        self.lbl_average_display_oper.config(text=f"{average_force} kg")
                    if hasattr(self, 'lbl_average_display_admin') and self.lbl_average_display_admin.winfo_exists():
                        self.lbl_average_display_admin.config(text=f"{average_force} kg")
                    
                    # Auto refrescar tabla de registros
                    self.refrescar_tabla_gui()
                else:
                    # El ciclo fue menor a 0.4s: se ignora como un "falso inicio" (colocación o ruido).
                    # Restauramos los displays visuales al último valor real registrado
                    if hasattr(self, 'lbl_average_display_oper') and self.lbl_average_display_oper.winfo_exists():
                        self.lbl_average_display_oper.config(text=f"{self.last_piece_force} kg" if self.last_piece_force > 0 else "-- kg")
                    if hasattr(self, 'lbl_average_display_admin') and self.lbl_average_display_admin.winfo_exists():
                        self.lbl_average_display_admin.config(text=f"{self.last_piece_force} kg" if self.last_piece_force > 0 else "-- kg")

                self.graph_draw_counter = 4  # Forzar redibujado de la curva final al cerrar el ciclo
                self.cycle_start_time = None

        self.piston_last_state = proceso_activo

        # Actualizar LEDs e Indicadores según el Perfil Activo
        if self.perfil_rol == "Administrador":
            def cambiar_color_led(canvas, led_id, estado):
                color = COLOR_OK if estado else COLOR_NOK
                borde = "#059669" if estado else "#DC2626"
                canvas.itemconfig(led_id, fill=color, outline=borde)

            cambiar_color_led(self.c_piston, self.led_piston, proceso_activo)
            cambiar_color_led(self.c_barrera, self.led_barrera, b_act)
            cambiar_color_led(self.c_emergencia, self.led_emergencia, not e_act)
            cambiar_color_led(self.c_inicio, self.led_inicio, i_act)
        else: # Perfil: Operador
            # 1. Estado del Proceso (Pistón)
            if proceso_activo:
                self.lbl_piston_status_oper.config(text="PROCESO ACTIVO", bg=COLOR_OK, fg="white")
            else:
                self.lbl_piston_status_oper.config(text="PROCESO INACTIVO", bg="#E2E8F0", fg=COLOR_TEXTO_SEC)
            
            # 2. Paro de Emergencia Gigante (Active-High: True = Pressed / False = OK)
            if not e_act:
                self.c_emergencia_oper.itemconfig(self.led_emergencia_oper, fill=COLOR_OK, outline="#059669")
                self.lbl_emergencia_text_oper.config(text="SISTEMA OK", fg=COLOR_OK)
            else:
                self.c_emergencia_oper.itemconfig(self.led_emergencia_oper, fill=COLOR_NOK, outline="#DC2626")
                self.lbl_emergencia_text_oper.config(text="⚠️ PARO DE EMERGENCIA ACTIVO ⚠️", fg=COLOR_NOK)

        # Actualizar Barra de Progreso del Ciclo en ambas vistas (Admin y Operador)
        v4_lim = max(0.1, v4)
        if proceso_activo and self.cycle_start_time is not None:
            elapsed = time.time() - self.cycle_start_time
            pct = min(elapsed / v4_lim, 1.0)
            status_text = f"Progreso del Ciclo: {elapsed:.1f} s / {v4:.1f} s"
        else:
            pct = 0.0
            status_text = f"Ciclo Inactivo (Límite: {v4:.1f} s)"
        
        # Actualizar textos de progreso
        if hasattr(self, 'lbl_progress_status_oper') and self.lbl_progress_status_oper.winfo_exists():
            self.lbl_progress_status_oper.config(text=status_text)
        if hasattr(self, 'lbl_progress_status_admin') and self.lbl_progress_status_admin.winfo_exists():
            self.lbl_progress_status_admin.config(text=status_text)
        
        # Actualizar anchos de la barra sin update_idletasks para evitar congelamiento
        try:
            if hasattr(self, 'frame_progress_oper') and self.frame_progress_oper.winfo_exists():
                total_w = self.frame_progress_oper.winfo_width()
                if total_w > 1:
                    target_w = int(total_w * pct)
                    self.bar_fill_oper.place_configure(width=target_w)
            if hasattr(self, 'frame_progress_admin') and self.frame_progress_admin.winfo_exists():
                total_w = self.frame_progress_admin.winfo_width()
                if total_w > 1:
                    target_w = int(total_w * pct)
                    self.bar_fill_admin.place_configure(width=target_w)
        except Exception:
            pass

        # Actualizar gráfica solo si el proceso está activo
        if proceso_activo:
            self.grafica_datos.append(v6_filtrado)
            self.datos_limite.append(v0)

        # Redibujar gráfica solo cada 5 ciclos (~500ms) para no bloquear el hilo GUI y mantener la respuesta táctil
        draw_cnt = getattr(self, 'graph_draw_counter', 0)
        draw_cnt += 1
        if draw_cnt >= 5:
            self.graph_draw_counter = 0
            self.line.set_xdata(list(range(len(self.grafica_datos))))
            self.line.set_ydata(list(self.grafica_datos))
            self.line_limite.set_xdata(list(range(len(self.datos_limite))))
            self.line_limite.set_ydata(list(self.datos_limite))
            self.ax.set_xlim(0, 50)

            max_val = max(self.grafica_datos) if self.grafica_datos else 0
            max_limit = max(max_val, v0)
            if max_limit > self.ax.get_ylim()[1]:
                self.ax.set_ylim(0, max_limit + (max_limit * 0.15))
            elif max_limit < self.ax.get_ylim()[1] * 0.5 and self.ax.get_ylim()[1] > 100:
                self.ax.set_ylim(0, max(100, max_limit + 15))
                
            self.canvas.draw()
        else:
            self.graph_draw_counter = draw_cnt

        # Log local en pantalla (cada 1s)
        self.tick_counter += 1
        if self.tick_counter >= 10:
            hora_actual = time.strftime("%H:%M:%S")
            self.listbox_log.insert(0, f"[{hora_actual}] Fuerza: {v6_filtrado} | P:{1 if p_act else 0} B:{1 if b_act else 0} E:{1 if e_act else 0} I:{1 if i_act else 0}")
            if self.listbox_log.size() > 100:
                self.listbox_log.delete(100, tk.END)
            self.tick_counter = 0
            self.refrescar_tabla_gui()

    # --- MANEJO DE HISTORIAL SQLITE EN TREEVIEW ---
    def refrescar_tabla_gui(self, fecha=None, h_ini=None, h_fin=None):
        if self.notebook.select() != self.notebook.tabs()[2]:
            return

        for item in self.tree_registros.get_children():
            self.tree_registros.delete(item)
            
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            query = "SELECT id, fecha, hora, valor_fuerza, resultado, limite_minimo, usuario FROM registros"
            params = []
            conditions = []
            
            if fecha:
                conditions.append("fecha = ?")
                params.append(fecha)
            if h_ini:
                conditions.append("hora >= ?")
                params.append(h_ini if len(h_ini) == 8 else f"{h_ini}:00")
            if h_fin:
                conditions.append("hora <= ?")
                params.append(h_fin if len(h_fin) == 8 else f"{h_fin}:59")
                
            if conditions:
                query += " WHERE " + " AND ".join(conditions)
                
            query += " ORDER BY id DESC LIMIT 100"
            
            cursor.execute(query, params)
            for reg in cursor.fetchall():
                res_val = reg[4] # "OK" o "NOK"
                tag = "RESULT_OK" if res_val == "OK" else "RESULT_NOK"
                self.tree_registros.insert("", "end", values=reg, tags=(tag,))
            conn.close()
            
            # Estilos del Treeview
            self.tree_registros.tag_configure("RESULT_OK", background="#E2F0D9", foreground="#2E7D32")
            self.tree_registros.tag_configure("RESULT_NOK", background="#FFEBEE", foreground="#C62828")
        except Exception as e:
            print(f"Error al refrescar tabla demo: {e}")

    def set_fecha_hoy(self):
        self.entry_fecha_filtro.delete(0, tk.END)
        self.entry_fecha_filtro.insert(0, datetime.now().strftime("%Y-%m-%d"))

    def aplicar_filtro_db(self):
        fecha = self.entry_fecha_filtro.get().strip()
        h_ini = self.combo_h_inicio.get()
        h_fin = self.combo_h_fin.get()
        if fecha:
            try:
                datetime.strptime(fecha, "%Y-%m-%d")
            except ValueError:
                messagebox.showerror("Error", "La fecha debe estar en formato AAAA-MM-DD")
                return
        self.refrescar_tabla_gui(fecha=fecha if fecha else None, h_ini=h_ini, h_fin=h_fin)

    def limpiar_filtro_db(self):
        self.set_fecha_hoy()
        self.combo_h_inicio.set("00:00")
        self.combo_h_fin.set("23:59")
        self.refrescar_tabla_gui()

    def abrir_calendario_modal(self):
        from calendar import monthcalendar, month_name
        
        now = datetime.now()
        curr_year = now.year
        curr_month = now.month
        
        val = self.entry_fecha_filtro.get().strip()
        if val:
            try:
                dt = datetime.strptime(val, "%Y-%m-%d")
                curr_year = dt.year
                curr_month = dt.month
            except ValueError:
                pass
                
        cal_win = tk.Toplevel(self.root)
        cal_win.title("Seleccionar Fecha")
        cal_win.configure(bg=COLOR_TARJETA)
        cal_win.resizable(False, False)
        cal_win.transient(self.root)
        cal_win.grab_set()
        
        header_frame = tk.Frame(cal_win, bg=COLOR_TARJETA)
        header_frame.pack(fill="x", pady=6)
        
        lbl_month_year = tk.Label(header_frame, text="", font=("Helvetica", 10, "bold"), fg=COLOR_TEXTO, bg=COLOR_TARJETA)
        
        grid_frame = tk.Frame(cal_win, bg=COLOR_TARJETA)
        grid_frame.pack(padx=12, pady=6)
        
        def render_month(y, m):
            nonlocal curr_year, curr_month
            curr_year, curr_month = y, m
            lbl_month_year.config(text=f"{month_name[m].upper()} {y}")
            
            for widget in grid_frame.winfo_children():
                widget.destroy()
                
            dias_semana = ["D", "L", "M", "M", "J", "V", "S"]
            for col_idx, d_name in enumerate(dias_semana):
                tk.Label(grid_frame, text=d_name, font=("Helvetica", 9, "bold"), fg=COLOR_TEXTO_SEC, bg=COLOR_TARJETA, width=4).grid(row=0, column=col_idx, pady=3)
                
            weeks = monthcalendar(y, m)
            for row_idx, week in enumerate(weeks, 1):
                for col_idx, day in enumerate(week):
                    if day == 0:
                        continue
                    def select_day(d=day):
                        date_str = f"{curr_year:04d}-{curr_month:02d}-{d:02d}"
                        self.entry_fecha_filtro.delete(0, tk.END)
                        self.entry_fecha_filtro.insert(0, date_str)
                        cal_win.destroy()
                        
                    bg_color = "#E2E8F0"
                    fg_color = COLOR_TEXTO
                    if val == f"{y:04d}-{m:02d}-{day:02d}":
                        bg_color = COLOR_VERDE_WKK
                        fg_color = "white"
                    elif day == now.day and y == now.year and m == now.month:
                        bg_color = "#C8E6C9"
                        
                    btn_day = tk.Button(grid_frame, text=str(day), font=("Helvetica", 9),
                                        fg=fg_color, bg=bg_color, bd=0, width=4, height=1, cursor="hand2",
                                        activebackground=COLOR_VERDE_OSCURO, activeforeground="white", command=select_day)
                    btn_day.grid(row=row_idx, column=col_idx, padx=1, pady=1)
                    
        def prev_month():
            y, m = curr_year, curr_month - 1
            if m == 0:
                y, m = y - 1, 12
            render_month(y, m)
            
        def next_month():
            y, m = curr_year, curr_month + 1
            if m == 13:
                y, m = y + 1, 1
            render_month(y, m)
            
        btn_prev = tk.Button(header_frame, text="◀", font=("Helvetica", 9), fg=COLOR_TEXTO_SEC, bg="#F1F3F5", bd=0, padx=6, command=prev_month)
        btn_prev.pack(side="left", padx=12)
        lbl_month_year.pack(side="left", expand=True)
        btn_next = tk.Button(header_frame, text="▶", font=("Helvetica", 9), fg=COLOR_TEXTO_SEC, bg="#F1F3F5", bd=0, padx=6, command=next_month)
        btn_next.pack(side="right", padx=12)
        
        render_month(curr_year, curr_month)

    def confirmar_y_eliminar_db(self):
        res = messagebox.askyesno(
            "Advertencia de Eliminación",
            "⚠️ ADVERTENCIA CRÍTICA ⚠️\n\n"
            "Esta acción eliminará de forma permanente TODOS los registros locales de esta máquina.\n\n"
            "Se recomienda realizar una exportación a Excel antes de continuar para evitar pérdida de información.\n\n"
            "¿Desea proceder?",
            parent=self.root
        )
        if not res:
            return

        # Pedir contraseña de Administrador
        pass_win = tk.Toplevel(self.root)
        pass_win.title("Confirmar Eliminación")
        pass_win.configure(bg=COLOR_TARJETA)
        pass_win.resizable(False, False)
        pass_win.transient(self.root)
        pass_win.grab_set()

        w_w, w_h = 320, 160
        s_w = pass_win.winfo_screenwidth()
        s_h = pass_win.winfo_screenheight()
        pos_x = (s_w - w_w) // 2
        pos_y = (s_h - w_h) // 2
        pass_win.geometry(f"{w_w}x{w_h}+{pos_x}+{pos_y}")

        tk.Label(pass_win, text="Contraseña de Administrador:", font=("Helvetica", 10, "bold"), fg=COLOR_TEXTO, bg=COLOR_TARJETA).pack(pady=(15, 8))
        entry_pass_del = tk.Entry(pass_win, font=("Helvetica", 12), show="*", justify="center", bg="#F8F9FA", fg=COLOR_TEXTO, highlightbackground=COLOR_BORDE, highlightthickness=1, relief="flat")
        entry_pass_del.pack(pady=5, padx=20, fill="x")
        entry_pass_del.focus()

        def ejecutar_eliminacion():
            pwd = entry_pass_del.get()
            if pwd == "admin":
                try:
                    conn = sqlite3.connect(DB_PATH)
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM registros")
                    conn.commit()
                    conn.close()

                    self.refrescar_tabla_gui()
                    self.listbox_log.delete(0, tk.END)
                    messagebox.showinfo("Éxito", "Todos los registros demo han sido eliminados.", parent=pass_win)
                    pass_win.destroy()
                except Exception as e:
                    messagebox.showerror("Error", f"No se pudo limpiar la base de datos:\n{e}", parent=pass_win)
            else:
                messagebox.showerror("Error de Acceso", "Contraseña incorrecta.", parent=pass_win)

        btn_frame = tk.Frame(pass_win, bg=COLOR_TARJETA)
        btn_frame.pack(fill="x", pady=12)

        tk.Button(btn_frame, text="Confirmar", font=("Helvetica", 9, "bold"), fg="white", bg=COLOR_VERDE_WKK, bd=0, padx=12, pady=6, command=ejecutar_eliminacion).pack(side="left", padx=(30, 10), expand=True, fill="x")
        tk.Button(btn_frame, text="Cancelar", font=("Helvetica", 9, "bold"), fg=COLOR_TEXTO_SEC, bg="#E2E8F0", bd=0, padx=12, pady=6, command=pass_win.destroy).pack(side="right", padx=(10, 30), expand=True, fill="x")

    # --- EXPORTACION A EXCEL ---
    def exportar_a_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Obtener datos actualmente cargados en base al filtro de interfaz
            fecha = self.entry_fecha_filtro.get().strip()
            h_ini = self.combo_h_inicio.get()
            h_fin = self.combo_h_fin.get()

            try:
                conn = sqlite3.connect(DB_PATH)
                cursor = conn.cursor()
                query = "SELECT id, fecha, hora, valor_fuerza, resultado, limite_minimo, usuario FROM registros"
                params = []
                conditions = []
                if fecha:
                    conditions.append("fecha = ?")
                    params.append(fecha)
                if h_ini:
                    conditions.append("hora >= ?")
                    params.append(h_ini if len(h_ini) == 8 else f"{h_ini}:00")
                if h_fin:
                    conditions.append("hora <= ?")
                    params.append(h_fin if len(h_fin) == 8 else f"{h_fin}:59")
                if conditions:
                    query += " WHERE " + " AND ".join(conditions)
                query += " ORDER BY id DESC"
                cursor.execute(query, params)
                registros = cursor.fetchall()
                conn.close()
            except Exception as e:
                messagebox.showerror("Error DB", f"No se pudo consultar la base de datos para exportar:\n{e}")
                return

            if not registros:
                messagebox.showinfo("Exportar Excel", "No hay registros disponibles bajo el filtro actual.")
                return

            # Crear libro de trabajo openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Historial de Prensa"
            ws.views.sheetView[0].showGridLines = True

            # Encabezados
            headers = [
                "ID", "Fecha", "Hora", "Fuerza Registrada (VW6)", 
                "Resultado (OK/NOK)", "Límite Mínimo (VW0)", "Usuario"
            ]

            # Estilos del Excel
            font_header = Font(name="Arial", bold=True, size=11, color="FFFFFF")
            fill_header = PatternFill(start_color="0E8A3E", end_color="0E8A3E", fill_type="solid") # WKK Verde
            center_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin', color='DDDDDD'),
                right=Side(style='thin', color='DDDDDD'),
                top=Side(style='thin', color='DDDDDD'),
                bottom=Side(style='thin', color='DDDDDD')
            )

            # Escribir encabezados
            ws.append(headers)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = center_align
                cell.border = thin_border

            # Escribir filas
            fill_ok = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
            fill_nok = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
            for row_idx, r_data in enumerate(registros, 2):
                for col_idx, val in enumerate(r_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.alignment = center_align
                    cell.border = thin_border
                    
                    if col_idx == 5:
                        if val == "OK":
                            cell.fill = fill_ok
                        elif val == "NOK":
                            cell.fill = fill_nok

            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            # En Windows dialog normal
            fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = filedialog.asksaveasfilename(
                parent=self.root,
                title="Guardar registros demo...",
                initialdir=SCRIPT_DIR,
                initialfile=f"registros_prensa_demo_{fecha_str}.xlsx",
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
            )

            if not filename:
                return

            try:
                wb.save(filename)
                messagebox.showinfo("Exportar Excel", f"Archivo exportado con éxito:\n{filename}")
            except Exception as e:
                messagebox.showerror("Error", f"Error al guardar el archivo:\n{e}")

        except ImportError:
            messagebox.showerror("Error de Dependencias", "Se requiere openpyxl instalado para exportar a Excel.\nEjecuta: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Error", f"Fallo al exportar:\n{e}")

    # --- MÉTODOS DE GESTIÓN DE USUARIOS ---
    def refrescar_usuarios_gui(self):
        for item in self.tree_usuarios.get_children():
            self.tree_usuarios.delete(item)
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT username, role FROM usuarios ORDER BY role ASC, username ASC")
            for row in cursor.fetchall():
                self.tree_usuarios.insert("", "end", values=row)
            conn.close()
        except Exception as e:
            print(f"Error al refrescar lista de usuarios demo: {e}")

    def on_usuario_select(self, event):
        selected = self.tree_usuarios.selection()
        if not selected:
            self.lbl_selected_user_name.config(text="Selecciona un usuario de la lista", fg=COLOR_TEXTO_SEC)
            return
        item = self.tree_usuarios.item(selected[0])
        username, role = item["values"]
        self.lbl_selected_user_name.config(text=f"Usuario: {username} ({role})", fg=COLOR_TEXTO)

    def crear_usuario_gui(self):
        username = self.entry_new_user.get().strip()
        password = self.entry_new_pass.get()
        role = self.combo_new_role.get()
        
        if not username:
            messagebox.showerror("Error", "El nombre de usuario es requerido.")
            return
            
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("INSERT INTO usuarios (username, password, role) VALUES (?, ?, ?)", (username, password, role))
            conn.commit()
            conn.close()
            
            messagebox.showinfo("Éxito", f"Usuario '{username}' creado exitosamente.")
            self.entry_new_user.delete(0, tk.END)
            self.entry_new_pass.delete(0, tk.END)
            self.refrescar_usuarios_gui()
        except sqlite3.IntegrityError:
            messagebox.showerror("Error", f"El usuario '{username}' ya existe.")
        except Exception as e:
            messagebox.showerror("Error", f"Error al crear usuario:\n{e}")

    def guardar_contrasena_gui(self):
        selected = self.tree_usuarios.selection()
        if not selected:
            messagebox.showwarning("Aviso", "Por favor selecciona un usuario de la lista.")
            return
            
        item = self.tree_usuarios.item(selected[0])
        username = item["values"][0]
        new_pass = self.entry_edit_pass.get()
        
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("UPDATE usuarios SET password = ? WHERE username = ?", (new_pass, username))
            conn.commit()
            conn.close()
            
            messagebox.showinfo("Éxito", f"Contraseña actualizada para el usuario '{username}'.")
            self.entry_edit_pass.delete(0, tk.END)
            self.refrescar_usuarios_gui()
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar contraseña:\n{e}")

    def eliminar_usuario_gui(self):
        selected = self.tree_usuarios.selection()
        if not selected:
            messagebox.showwarning("Aviso", "Por favor selecciona un usuario de la lista.")
            return
            
        item = self.tree_usuarios.item(selected[0])
        username, role = item["values"]
        
        if username == self.perfil_actual:
            messagebox.showerror("Error", "No puedes eliminar al usuario activo con el que iniciaste sesión.")
            return
            
        if username == "admin":
            messagebox.showerror("Error", "No se puede eliminar el usuario administrador maestro 'admin'.")
            return
            
        res = messagebox.askyesno("Confirmar", f"¿Estás seguro de que deseas eliminar al usuario '{username}'?")
        if not res:
            return
            
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM usuarios WHERE username = ?", (username,))
            conn.commit()
            conn.close()
            
            messagebox.showinfo("Éxito", f"Usuario '{username}' eliminado.")
            self.lbl_selected_user_name.config(text="Selecciona un usuario de la lista", fg=COLOR_TEXTO_SEC)
            self.refrescar_usuarios_gui()
        except Exception as e:
            messagebox.showerror("Error", f"Error al eliminar usuario:\n{e}")

    # --- CIERRE DE APLICACIÓN ---
    def on_closing(self):
        self.running_loop = False  
        time.sleep(0.2)
        try:
            self.plc_client.disconnect()
        except Exception:
            pass
        self.root.quit()
        self.root.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app = LogoHMI(root)
    root.mainloop()
